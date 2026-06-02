# -*- coding: utf-8 -*-
"""Exportacao CSV/XLSX do modulo de relatorios com metadados de filtros."""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Sequence

from application.reports_contracts import ReportsDetailDTO, ReportsFilterDTO

_HEADERS = ("Data", "Exame", "Amostra", "Resultado", "Status GAL", "Analista", "Kit", "Lote")
_COL_WIDTHS = (12, 30, 14, 18, 14, 16, 14, 16)


def _detail_to_row(d: ReportsDetailDTO) -> tuple:
    return (
        str(d.data_exame),
        d.exame,
        d.amostra_codigo,
        d.resultado_geral,
        d.status_gal,
        d.analista or "",
        d.kit or "",
        d.lote or "",
    )


def _build_metadata(
    filters: ReportsFilterDTO, usuario: str, total: int
) -> list[tuple[str, str]]:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return [
        ("Gerado em", now),
        ("Periodo", f"{filters.data_inicio} a {filters.data_fim}"),
        ("Usuario", usuario),
        ("Exames", ", ".join(filters.exames) or "Todos"),
        ("Positividade", ", ".join(filters.positividade) or "Todos"),
        ("Status GAL", ", ".join(filters.status_gal) or "Todos"),
        ("Analista", ", ".join(filters.analistas) or "Todos"),
        ("Kit", ", ".join(filters.kits) or "Todos"),
        ("Lote", ", ".join(filters.lotes) or "Todos"),
        ("Total exportado", str(total)),
    ]


def export_csv(
    detalhes: Sequence[ReportsDetailDTO],
    filters: ReportsFilterDTO,
    usuario: str,
    filepath: Path,
) -> None:
    """Grava CSV com metadados de cabecalho (comentarios #) e dados tabulares."""
    meta = _build_metadata(filters, usuario, len(detalhes))
    with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        for key, value in meta:
            writer.writerow([f"# {key}: {value}"])
        writer.writerow([])
        writer.writerow(_HEADERS)
        for d in detalhes:
            writer.writerow(_detail_to_row(d))


def export_xlsx(
    detalhes: Sequence[ReportsDetailDTO],
    filters: ReportsFilterDTO,
    usuario: str,
    filepath: Path,
) -> None:
    """Grava XLSX com sheet 'Dados' e sheet 'Metadados'."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()

    # ---- Sheet Dados ----
    ws = wb.active
    ws.title = "Dados"

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1A5276")
    hdr_align = Alignment(horizontal="center")

    for col, text in enumerate(_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align

    for row, d in enumerate(detalhes, 2):
        for col, value in enumerate(_detail_to_row(d), 1):
            ws.cell(row=row, column=col, value=value)

    from openpyxl.utils import get_column_letter
    for col, width in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"

    # ---- Sheet Metadados ----
    wm = wb.create_sheet("Metadados")
    bold = Font(bold=True)
    wm.cell(row=1, column=1, value="Campo").font = bold
    wm.cell(row=1, column=2, value="Valor").font = bold
    for row, (key, value) in enumerate(_build_metadata(filters, usuario, len(detalhes)), 2):
        wm.cell(row=row, column=1, value=key)
        wm.cell(row=row, column=2, value=value)
    wm.column_dimensions["A"].width = 20
    wm.column_dimensions["B"].width = 55

    wb.save(filepath)
