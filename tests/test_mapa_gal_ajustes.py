# -*- coding: utf-8 -*-
"""Guardioes dos ajustes de Mapa da Placa (.xlsx) e do loteKit no CSV GAL.

Cobre:
  - loteKit do CSV GAL recebe o valor obrigatorio "Lote / Kit".
  - Cabecalho linha 2: info/Operador em A2:Q2 e validacao em R2:X2.
  - Rodape: CN em A54:O54, CP em A55:O55 com todos os alvos+RP (ALVO - CT),
    bloco vazio P54:X56, todos com borda externa preta.
  - Cores de controle CN/CP por validade no plate viewer.
"""

from types import SimpleNamespace

import pandas as pd
from openpyxl import Workbook


# ---------------------------------------------------------------------------
# Item 2 — loteKit no CSV GAL
# ---------------------------------------------------------------------------

def test_loteKit_recebe_lote_kit():
    from exportacao.gal_formatter import _build_base_dataframe

    cfg = SimpleNamespace(kit_codigo="1175")
    df_out = _build_base_dataframe(
        cfg=cfg,
        cod_col=pd.Series(["101", "102"]),
        exam_value="VRSRT",
        panel_value="12",
        lote_kit="LOTE-ABC",
    )
    assert list(df_out["loteKit"]) == ["LOTE-ABC", "LOTE-ABC"]


def test_loteKit_default_vazio_preserva_retrocompat():
    from exportacao.gal_formatter import _build_base_dataframe

    cfg = SimpleNamespace(kit_codigo="1175")
    df_out = _build_base_dataframe(
        cfg=cfg,
        cod_col=pd.Series(["101"]),
        exam_value="VRSRT",
        panel_value="12",
    )
    assert list(df_out["loteKit"]) == [""]


# ---------------------------------------------------------------------------
# Item 3 — Cabecalho linha 2 (A2:Q2 + R2:X2)
# ---------------------------------------------------------------------------

def _mapa_minimo(placa_ok=True):
    from domain.mapa_placa_layout import MapaPlaca

    return MapaPlaca(
        nome_exame="ZDC BioManguinhos",
        nome_placa="placa_680",
        placa_ok=placa_ok,
        blocos_por_linha=6,
        linhas_blocos=8,
        blocos=(),
        controles=(),
    )


def _merges(ws):
    return {str(r) for r in ws.merged_cells.ranges}


def test_cabecalho_linha2_merges():
    from exportacao.mapa_placa_exporter import _escrever_cabecalho

    wb = Workbook()
    ws = wb.active
    _escrever_cabecalho(ws, _mapa_minimo(), total_cols_excel=24, nome_operador="fulano")

    merges = _merges(ws)
    assert "A1:X1" in merges  # nome do exame
    assert "A2:Q2" in merges  # info + Operador
    assert "R2:X2" in merges  # validacao da placa
    assert "Operador: fulano" in str(ws["A2"].value)
    assert "ZDC BIOMANGUINHOS" in str(ws["A1"].value)
    # Item 4: fonte da A2:Q2 = 12.
    assert int(ws["A2"].font.size) == 12


def test_cabecalho_status_placa_ok_e_invalida():
    from exportacao.mapa_placa_exporter import _escrever_cabecalho

    # placa_valida=True -> "PLACA OK"
    wb = Workbook()
    ws = wb.active
    _escrever_cabecalho(ws, _mapa_minimo(placa_ok=False), 24, "op", placa_valida=True)
    assert ws["R2"].value == "PLACA OK"

    # placa_valida=False -> "PLACA INVÁLIDA"
    wb2 = Workbook()
    ws2 = wb2.active
    _escrever_cabecalho(ws2, _mapa_minimo(placa_ok=True), 24, "op", placa_valida=False)
    assert ws2["R2"].value == "PLACA INVÁLIDA"


def test_placa_valida_do_df_le_status_placa():
    from exportacao.mapa_placa_exporter import _placa_valida_do_df

    df_ok = pd.DataFrame([{"Amostra": "CN", "Status_Placa": "Válida"}])
    df_inv = pd.DataFrame([{"Amostra": "CN", "Status_Placa": "Inválida (CN incorreto)"}])
    df_indef = pd.DataFrame([{"Amostra": "CN", "Status_Placa": "Indefinido"}])
    df_sem = pd.DataFrame([{"Amostra": "CN"}])

    assert _placa_valida_do_df(df_ok) is True
    assert _placa_valida_do_df(df_inv) is False
    assert _placa_valida_do_df(df_indef) is False
    assert _placa_valida_do_df(df_sem) is None


def _df_controles(cp_sc2="Detectável"):
    return pd.DataFrame(
        [
            {"Amostra": "CN", "Res_RP_1": "Válido", "Res_RP_2": "Válido", "Res_SC2": "Não Detectável"},
            {"Amostra": "CP", "Res_RP_1": "Válido", "Res_RP_2": "Válido", "Res_SC2": cp_sc2},
        ]
    )


def test_status_placa_reflete_edicao_de_ct():
    """Apos editar CT (=> Res_ muda), a validade recalculada acompanha.

    Garante a Opcao A: o Mapa Definitivo (que le Status_Placa) reflete edicoes.
    """
    from services.analysis.analysis_service import _avaliar_status_placa_vectorized
    from exportacao.mapa_placa_exporter import _placa_valida_do_df

    alvos_cols_res = ["Res_SC2"]

    df_ok = _df_controles(cp_sc2="Detectável")
    df_ok["Status_Placa"] = _avaliar_status_placa_vectorized(df_ok, alvos_cols_res)
    assert _placa_valida_do_df(df_ok) is True

    # Edicao que torna o CP invalido (alvo deixou de detectar).
    df_inv = _df_controles(cp_sc2="Não Detectável")
    df_inv["Status_Placa"] = _avaliar_status_placa_vectorized(df_inv, alvos_cols_res)
    assert _placa_valida_do_df(df_inv) is False


# ---------------------------------------------------------------------------
# Item 4/5 — Rodape de controles + bloco vazio
# ---------------------------------------------------------------------------

def test_rodape_controles_geometria_texto_e_bordas():
    from exportacao.mapa_placa_exporter import _escrever_rodape_controles

    wb = Workbook()
    ws = wb.active
    detalhe = {
        "CN": "ZK - 30,12   RP1 - 28,50",
        "CP": "ZK - 22,00   RP1 - 27,10",
    }
    _escrever_rodape_controles(ws, None, 53, 24, detalhe)

    merges = _merges(ws)
    assert "A54:O54" in merges  # CN
    assert "A55:O55" in merges  # CP
    assert "P54:X56" in merges  # bloco vazio

    assert ws["A54"].value == "CN: ZK - 30,12   RP1 - 28,50"
    assert ws["A55"].value.startswith("CP: ")
    assert ws["P54"].value is None  # bloco vazio

    # Bordas externas pretas: cantos do retangulo
    assert ws["A54"].border.left.style is not None
    assert ws["A54"].border.top.style is not None
    assert ws["P54"].border.top.style is not None
    assert ws["P54"].border.left.style is not None
    assert ws["X56"].border.bottom.style is not None
    assert ws["X56"].border.right.style is not None


def test_coletar_alvos_controles_inclui_rp_e_omite_sem_ct():
    from exportacao.mapa_placa_exporter import _coletar_alvos_controles

    df = pd.DataFrame(
        [
            {"Amostra": "CN", "CT_ZK": "", "CT_RP1": "28,50", "CT_DEN1": ""},
            {"Amostra": "CP", "CT_ZK": "22,0", "CT_RP1": "27,1", "CT_DEN1": ""},
            {"Amostra": "101", "CT_ZK": "30,0", "CT_RP1": "28,0", "CT_DEN1": ""},
        ]
    )
    detalhe = _coletar_alvos_controles(df)

    # CN: ZK sem CT (omitido), RP1 presente. Formato compacto "ALVO-CT" unido por "| ".
    assert detalhe["CN"] == "RP1-28,50"
    assert "ZK" not in detalhe["CN"]
    # CP: ZK e RP1 presentes, separados por "| ".
    assert detalhe["CP"] == "ZK-22,00| RP1-27,10"


# ---------------------------------------------------------------------------
# Item 1 — Cor CN/CP por validade (plate viewer)
# ---------------------------------------------------------------------------

def test_status_colors_cn_cp_trocadas():
    from ui.components.plate_viewer import (
        STATUS_COLORS,
        CONTROL_CN,
        CONTROL_CP,
    )
    from ui.theme.design_tokens import SemanticColors

    # CN agora usa a cor verde (antiga do CP) e vice-versa.
    assert STATUS_COLORS[CONTROL_CN] == SemanticColors.CONTROLE_CP
    assert STATUS_COLORS[CONTROL_CP] == SemanticColors.CONTROLE_CN


def test_controle_valido_regras():
    from ui.components.plate_viewer import controle_valido

    cn_alvos = {"ZK": SimpleNamespace(result="Não Detectável"), "RP1": SimpleNamespace(result="Detectável")}
    cp_alvos = {"ZK": SimpleNamespace(result="Detectável"), "RP1": SimpleNamespace(result="Detectável")}

    # CN valido = sem detectavel analitico (RP nao conta).
    assert controle_valido(cn_alvos, "CN") is True
    # CN com alvo analitico detectavel = invalido.
    assert controle_valido(cp_alvos, "CN") is False
    # CP valido = ao menos um detectavel analitico.
    assert controle_valido(cp_alvos, "CP") is True
    assert controle_valido(cn_alvos, "CP") is False
