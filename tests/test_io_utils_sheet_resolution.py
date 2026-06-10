# -*- coding: utf-8 -*-
"""Guardioes da resolucao de aba na leitura generica de resultados.

Regressao do bug: planilha do 7500 (sem aba "Results") falhava com
`Worksheet named 'Results' not found` -> df None -> '.shape' em None.
"""

import openpyxl
import pandas as pd
import pytest

from utils.io_utils import _resolver_sheet_resultados, read_data_with_auto_detection


def _criar_xlsx(path, sheet_name, *, metadados=2):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    # Linhas de metadados antes do cabecalho (como nos exports reais).
    for i in range(metadados):
        ws.append([f"meta {i}", "", "", ""])
    ws.append(["Well", "Sample Name", "Target Name", "Ct"])
    ws.append(["A1", "101", "SC2", "30,12"])
    ws.append(["A1", "101", "RP", "28,50"])
    wb.save(path)
    return path


def test_resolver_sheet_prefere_results(tmp_path):
    p = _criar_xlsx(tmp_path / "quant.xlsx", "Results")
    assert _resolver_sheet_resultados(str(p)) == "Results"


def test_resolver_sheet_fallback_para_primeira_aba(tmp_path):
    # 7500: aba com outro nome (sem "Results").
    p = _criar_xlsx(tmp_path / "p7500.xlsx", "20250718 BM VR1_VR2 PL 5")
    assert _resolver_sheet_resultados(str(p)) == "20250718 BM VR1_VR2 PL 5"


def test_resolver_sheet_ignora_aba_de_extracao(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "PLANILHA EXTRAÇÃO"
    ws2 = wb.create_sheet("Dados 7500")
    ws2.append(["Well", "Sample Name", "Target Name", "Ct"])
    ws2.append(["A1", "101", "SC2", "30,1"])
    p = tmp_path / "extra.xlsx"
    wb.save(p)
    assert _resolver_sheet_resultados(str(p)) == "Dados 7500"


def test_read_data_le_planilha_7500_sem_results(tmp_path):
    p = _criar_xlsx(tmp_path / "p7500.xlsx", "Sheet1")
    df = read_data_with_auto_detection(str(p))
    assert df is not None
    cols = [str(c) for c in df.columns]
    assert any("Well" in c for c in cols)
    assert any("Sample" in c for c in cols)
    assert any("Target" in c for c in cols)


def test_read_data_results_ainda_funciona(tmp_path):
    p = _criar_xlsx(tmp_path / "quant.xlsx", "Results")
    df = read_data_with_auto_detection(str(p))
    assert df is not None
    assert any("Target" in str(c) for c in df.columns)
