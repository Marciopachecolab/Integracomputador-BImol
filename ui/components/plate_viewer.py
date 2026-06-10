"""
Visualização interativa da placa (CustomTkinter) consumindo dados em memória
vindo da análise (df_final/df_norm). Inspirado em services/teste_plate_viewer_historico_ctk.py,
mas sem ler CSV e sem exportar XLS automaticamente.
"""

from __future__ import annotations

import math
import os
import unicodedata
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional

import customtkinter as ctk
from config.enums import ResultStatus
import pandas as pd
import tkinter as tk
from tkinter import ttk
from utils.csv_lock import CSVFileLock
from utils.text_normalizer import repair_mojibake_text
from utils.text_result_classifier import classify_result_text

# Tenta importar a configuração de exame, se falhar usa None
try:
    from services.exam_registry import get_exam_cfg
except ImportError:
    get_exam_cfg = None

# ---------------------------------------------------------------------------
# Aparência / Cores
# ---------------------------------------------------------------------------

NEGATIVE = "NEGATIVE"
POSITIVE = "POSITIVE"
INDETERMINADO = "INDETERMINADO"
INCONCLUSIVE = INDETERMINADO
INVALID = "INVALID"
CONTROL_CN = "CONTROL_CN"
CONTROL_CP = "CONTROL_CP"
EMPTY = "EMPTY"
FILLED = "FILLED"

from ui.theme.design_tokens import SemanticColors

STATUS_COLORS = {
    NEGATIVE: SemanticColors.NAO_DETECTAVEL,
    POSITIVE: SemanticColors.DETECTADO,
    INDETERMINADO: SemanticColors.INCONCLUSIVO,
    INVALID: SemanticColors.INVALIDO,
    # Cores invertidas (CN=verde, CP=vermelho) quando o controle e valido.
    CONTROL_CN: SemanticColors.CONTROLE_CP,  # verde
    CONTROL_CP: SemanticColors.CONTROLE_CN,  # vermelho
    EMPTY: SemanticColors.EMPTY,
}

# Cores para diferentes tamanhos de grupos (exames de 48, 32, 24 testes)
# Cores atualizadas conforme o prompt
from ui.theme.design_tokens import GroupColors

GROUP_COLORS = {
    2: GroupColors.PAIR,
    3: GroupColors.TRIO,
    4: GroupColors.QUARTET,
}

ROW_LABELS = ["A", "B", "C", "D", "E", "F", "G", "H"]
COL_LABELS = [str(i) for i in range(1, 13)]

# ---------------------------------------------------------------------------
# Modelos
# ---------------------------------------------------------------------------


@dataclass
class TargetResult:
    result: str = ""  # "Det", "ND", "Inc", "Inv"...
    ct: Optional[float] = None
    resultado_manual: bool = False  # True quando resultado foi definido manualmente pelo usuario


@dataclass
class WellData:
    row_label: str
    col_label: str
    well_id: str
    sample_id: Optional[str] = None
    code: Optional[str] = None
    status: str = EMPTY
    is_control: bool = False
    targets: Dict[str, TargetResult] = field(default_factory=dict)
    metadata: Dict[str, Any] = field(default_factory=dict)
    # Campos para grupos de poços
    paired_wells: List[str] = field(default_factory=list)
    is_grouped: bool = False
    group_id: Optional[str] = None
    group_size: int = 1
    group_position: int = 0
    # âœ… Bug #15: Adicionar atributo dinâmico usado em linhas 378, 574-586
    pair_group_id: Optional[str] = None  # ID do grupo de pares (legacy)
    @property
    def poco(self) -> str:
        return self.well_id


@dataclass
class WellResult:
    """Payload de exportacao para mapa de placa."""

    well: str
    sample_code: str
    ct_target: Optional[float]
    ct_rp: Optional[float]
    result: str
    is_control: bool
    control_type: Optional[str] = None


class Well(WellData):
    """
    Compatibilidade com testes legados que esperam Well(row='A', col=1).
    """

    def __init__(self, row: str, col: int, **kwargs: Any):
        row_label = str(row).upper()
        col_label = str(col)
        well_id = f"{row_label}{col_label}"
        super().__init__(row_label=row_label, col_label=col_label, well_id=well_id, **kwargs)



# FASE 2.1: Função de normalização de well IDs
def normalize_well_id(well_id: str) -> str:
    """
    Normaliza ID de poço removendo zeros Ã  esquerda.
    
    Converte IDs como 'A01' para 'A1', mantendo 'A10' como 'A10'.
    Ãštil para compatibilidade entre diferentes formatos de exportação de equipamentos.
    
    Args:
        well_id: ID do poço no formato 'A01', 'H12', etc.
        
    Returns:
        str: Well ID normalizado sem zeros Ã  esquerda
        
    Examples:
        >>> normalize_well_id('A01')
        'A1'
        >>> normalize_well_id('H12')
        'H12'
        >>> normalize_well_id('B03')
        'B3'
        
    Raises:
        ValueError: Se well_id não for uma string válida
    """
    if not well_id or not isinstance(well_id, str):
        raise ValueError(f"well_id deve ser string não-vazia, recebido: {well_id}")
    
    import re
    well_id_str = str(well_id).strip().upper()
    match = re.match(r'([A-H])0*(\d+)', well_id_str)
    if match:
        return f"{match.group(1)}{match.group(2)}"
    return well_id_str


class WellDict(dict):
    """Dicionario de wells que itera sobre valores (compatibilidade testes)."""

    def __iter__(self):
        return iter(self.values())


class PlateModel:
    def __init__(self, rows: int = 8, cols: int = 12) -> None:
        self.wells: WellDict[str, WellData] = WellDict()
        self.group_dict: Dict[str, List[str]] = {}  # Novo: dicionário de grupos
        self.pair_groups: Dict[str, List[str]] = {}  # Legado: compatibilidade
        self.exam_type: str = "96"  # Tipo de exame: 96, 48, 32, 24 testes
        self.requires_group_frames: bool = False
        self.group_size: int = 1  # Mantido para compatibilidade
        self.exam_cfg: Optional[Any] = None
        self.include_resultado_cols: bool = False
        self.include_resultado_geral: bool = False
        self.rows = rows
        self.cols = cols

    # ------------------ construção a partir de df ------------------ #
    
    @classmethod
    def from_df(
        cls,
        df_final: pd.DataFrame,
        group_size: Optional[int] = None,
        exame: Optional[str] = None,
    ) -> "PlateModel":
        """
        Constrói o modelo de placa a partir dos dados da corrida.
        """
        model = cls()
        # se foi passado nome do exame, carregue a configuração correspondente
        if exame and get_exam_cfg:
            try:
                model.exam_cfg = get_exam_cfg(exame)
            except Exception:
                model.exam_cfg = None

        # determina tamanho de grupo preferindo valor do registry se não fornecido
        if group_size is None and getattr(model, "exam_cfg", None) is not None:
            try:
                group_size = model.exam_cfg.bloco_size()
            except Exception:
                group_size = None
        model.group_size = group_size or 1
        
        # Validação em duas etapas para evitar AttributeError
        if df_final is None:
            return model
        if df_final is None or df_final.empty:
            return model
        
        # Carrega config do registry se exame foi fornecido
        exam_cfg = model.exam_cfg

        df_use = df_final.copy()

        # Mapeia nomes de colunas para upper/lower
        cols_upper = {c: str(c).upper() for c in df_use.columns}
        cols_lower = {str(c).lower(): c for c in df_use.columns}

        # Se não houver coluna de poço estilo df_final, mas houver WELL, converte df_norm -> df_final-like
        has_poco = any(cu in {"POÃ‡O", "POCO"} for cu in cols_upper.values())
        has_well = any(cu == "WELL" or cu == "WELL_ID" for cu in cols_upper.values())
        if not has_poco and has_well:
            df_use = cls._convert_df_norm(df_use)
            cols_upper = {c: str(c).upper() for c in df_use.columns}
            cols_lower = {str(c).lower(): c for c in df_use.columns}
        
        upper_cols = [str(c).upper() for c in df_use.columns]
        model.include_resultado_cols = any(
            c.startswith("RESULTADO_") and c != "RESULTADO_GERAL" for c in upper_cols
        )
        model.include_resultado_geral = "RESULTADO_GERAL" in upper_cols

        # ------------------ identificação de colunas básicas ------------------ #
        # Poço / Poco
        poco_col = None
        for key in ["poÃ§o", "poco", "well", "well_id"]:
            c = cols_lower.get(key)
            if c:
                poco_col = c
                break
        if poco_col is None:
            # sem coluna de poço não há como desenhar a placa
            return model

        # Amostra
        sample_col = None
        for key in ["amostra", "sample", "samplename", "sample_name"]:
            c = cols_lower.get(key)
            if c:
                sample_col = c
                break

        # Código
        code_col = None
        for key in ["cÃ³digo", "codigo", "code"]:
            c = cols_lower.get(key)
            if c:
                code_col = c
                break

        # Descobre targets a partir das colunas Resultado_*
        targets: List[str] = []
        for c in df_use.columns:
            cu = cols_upper[c]
            if cu.startswith("RESULTADO_"):
                alvo = cu.replace("RESULTADO_", "").strip()
                if alvo and alvo not in targets:
                    targets.append(alvo)
            elif cu.startswith("RES_"):  # âœ… ETAPA 1: Formato alternativo Res_ADV, Res_HMPV
                alvo = cu.replace("RES_", "").strip()
                if alvo and alvo not in targets:
                    targets.append(alvo)
            elif cu.endswith(" - R") or cu.endswith("- R"):
                base = cu.replace(" - R", "").replace("- R", "").strip()
                if base and base not in targets:
                    targets.append(base)
        
        # FASE 1 - LOG CRÃTICO: Verificar detecção de targets
        from utils.logger import registrar_log
        registrar_log("PlateModel", f"Targets detectados: {targets}", "INFO")
        if not targets:
            registrar_log("PlateModel", "AVISO: Nenhum target detectado - wells ficarão sem alvos (cores brancas)", "WARNING")
        
        # Função de normalização para matching alvo vs coluna de CT
        def _norm_key(txt: str) -> str:
            # Remove parênteses antes de filtrar (para suportar C(t))
            txt_clean = str(txt).replace("(", "").replace(")", "")
            return "".join(ch for ch in txt_clean.upper() if ch.isalnum())

        def _find_ct_column_for_target(alvo: str) -> Optional[str]:
            """
            Procura coluna de CT compatível com o alvo.
            """
            target_key = _norm_key(alvo)
            for c in df_use.columns:
                cu = str(c).upper()
                # ignora colunas de resultado
                if cu.startswith("RESULTADO_") or cu.endswith(" - R") or cu.endswith("- R"):
                    continue
                base = cu
                if " - CT" in base:
                    base = base.split(" - CT")[0]
                if base.startswith("CT_"):
                    base = base[3:]
                col_key = _norm_key(base)
                if col_key == target_key:
                    return c
            return None

        # Deduz tamanho de agrupamento se não vier
        model.group_size = group_size or cls._infer_group_size(df_use)

        # ------------------ construção dos poços ------------------ #
        for idx, row in df_use.iterrows():
            poco_raw = str(
                row.get("Poco", "")
                or row.get("POCO", "")
                or row.get(cols_lower.get("poco", ""), "")
                or row.get(poco_col, "")
            ).strip()
            if not poco_raw:
                continue

            # suporta múltiplos poços no mesmo registro (ex.: "A01+B01")
            pocos = [p.strip() for p in poco_raw.split("+") if p.strip()]
            if not pocos:
                continue

            sample = None
            if sample_col:
                sample = str(row.get(sample_col, "")).strip() or None

            # Detecta código preferindo campos usuais
            code = None
            code_candidates = []
            if code_col:
                code_candidates.append(row.get(code_col, ""))
            code_candidates.extend(
                [
                    row.get("Codigo", ""),
                    row.get("CÃ“DIGO", ""),
                    row.get("Cï¿½digo", ""),  # encoding antigo
                    row.get("CODE", ""),
                    sample,
                ]
            )
            for cval in code_candidates:
                s = str(cval).strip()
                if s:
                    code = s
                    break
            if code is None:
                code = sample

            # Resultados por alvo e CT
            target_data: Dict[str, TargetResult] = {}

            # Primeiro, CTs de RP (RP, RP_1, RP_2, ...)
            for c in df_use.columns:
                cu = cols_upper[c]
                if cu == "RP" or cu.startswith("RP_"):
                    try:
                        ct_val = row.get(c, None)
                        if ct_val is not None and str(ct_val).strip() != "":
                            ct_val = float(str(ct_val).replace(",", "."))
                        else:
                            ct_val = None
                    except Exception:
                        ct_val = None
                    target_data[cu] = TargetResult("", ct_val)

            # Depois, alvos analíticos principais
            for alvo in targets:
                # resultado qualitativo - tentar múltiplos formatos
                res_val = ""
                
                # Formato 1: "Resultado_SC2"
                res_col = f"Resultado_{alvo}"
                if res_col in row.index:
                    res_val = str(row.get(res_col, "")).strip()
                
                # Formato 1.5: "Res_ADV" (formato usado pelo df_analise normalizado)
                if not res_val:
                    res_col_short = f"Res_{alvo}"
                    if res_col_short in row.index:
                        res_val = str(row.get(res_col_short, "")).strip()
                
                # Formato 2: "SC2 - R"
                if not res_val:
                    alt_col = f"{alvo} - R"
                    if alt_col in row.index:
                        res_val = str(row.get(alt_col, "")).strip()
                
                # Formato 3: Coluna com nome do alvo contendo resultado completo
                if not res_val:
                    if alvo in row.index:
                        res_val = str(row.get(alvo, "")).strip()
                
                norm_res = normalize_result(str(res_val))

                # CT associado
                ct_col = _find_ct_column_for_target(alvo)
                ct_val = None
                if ct_col is not None and ct_col in row:
                    try:
                        raw_ct = row.get(ct_col, None)
                        if raw_ct is not None and str(raw_ct).strip() != "":
                            ct_val = float(str(raw_ct).replace(",", "."))
                    except Exception:
                        ct_val = None

                target_data[alvo] = TargetResult(norm_res, ct_val)
                
                # DEBUG: Log de valores para diagnóstico de cores
                if idx < 3:  # Apenas primeiras 3 linhas para não poluir log
                    registrar_log("PlateModel", f"[DEBUG] Row {idx} - Alvo={alvo}, res_col_in_row={res_col in row.index}, res_val='{res_val}' -> norm_res='{norm_res}'", "DEBUG")

            # Preenche wells (poços) e grupos
            normalized_pocos = []
            for poco in pocos:
                if len(poco) < 2:
                    continue
                row_label = poco[0].upper()
                col_label = poco[1:]
                try:
                    col_idx = int(col_label)
                    normalized_well = f"{row_label}{col_idx:02d}"
                    normalized_pocos.append(normalized_well)
                except Exception:
                    continue
            
            for idx_poco, well_id in enumerate(normalized_pocos):
                row_label = well_id[0]
                col_label = well_id[1:]

                if well_id not in model.wells:
                    wd = WellData(
                        row_label=row_label,
                        col_label=col_label,
                        well_id=well_id,
                        sample_id=sample or "",
                        code=code or "",
                        status=EMPTY,
                        is_control=False,
                        targets={},
                        metadata={},
                        paired_wells=[],
                        is_grouped=False,
                        group_id=None,
                        group_size=1,
                        group_position=0,
                    )
                    model.wells[well_id] = wd
                else:
                    wd = model.wells[well_id]
                    # Atualiza sample_id e code se ainda não tiverem sido definidos
                    if not wd.sample_id and sample:
                        wd.sample_id = sample
                    if not wd.code and code:
                        wd.code = code

                # merges targets (caso haja múltiplas linhas para o mesmo poço)
                for alvo, tr in target_data.items():
                    if alvo in wd.targets:
                        # mantém primeiro resultado, apenas atualiza CT se ainda não houver
                        if wd.targets[alvo].ct is None and tr.ct is not None:
                            wd.targets[alvo].ct = tr.ct
                    else:
                        wd.targets[alvo] = tr

                # grupos (pares/trios/quartetos) - sistema completo
                if len(normalized_pocos) > 1:
                    sorted_pocos = sorted(normalized_pocos)
                    group_id = "+".join(sorted_pocos)
                    wd.is_grouped = True
                    wd.group_id = group_id
                    wd.group_size = len(normalized_pocos)
                    wd.group_position = idx_poco
                    wd.paired_wells = [p for p in normalized_pocos if p != well_id]
                    # Adiciona ao group_dict
                    if group_id not in model.group_dict:
                        model.group_dict[group_id] = []
                    if well_id not in model.group_dict[group_id]:
                        model.group_dict[group_id].append(well_id)
                    # Mantém compatibilidade com sistema legado
                    wd.pair_group_id = group_id
                    model.pair_groups.setdefault(group_id, []).append(well_id)

                # Detecta controles usando cfg do exame quando disponível
                try:
                    ctrl = PlateModel._detect_control(wd.sample_id, wd.code, exam_cfg)
                    if ctrl:
                        wd.is_control = True
                        wd.metadata["control_type"] = ctrl
                    else:
                        wd.is_control = False
                        wd.metadata.pop("control_type", None)
                except Exception:
                    wd.is_control = False

                model._recompute_status(wd)
                model.wells[well_id] = wd

        # Define tamanho de bloco: prioritário é group_size, depois config do exame, por fim inferência
        if group_size:
            model.group_size = group_size
        elif exam_cfg:
            try:
                model.group_size = exam_cfg.bloco_size()
            except Exception:
                model.group_size = cls._infer_group_size(df_use)
        else:
            model.group_size = cls._infer_group_size(df_use)
        
        # Armazena config para uso em _recompute_status
        model.exam_cfg = exam_cfg
        
        # Determina tipo de exame e se requer frames de grupo
        model._determine_exam_type()
        model._determine_group_frame_requirement()
        
        return model
    @staticmethod
    def _convert_df_norm(df_norm: pd.DataFrame) -> pd.DataFrame:
        """
        Converte df_norm (linha por poço/target) em um df_final-like por poço,
        criando colunas Resultado_<ALVO> e CT_<ALVO>.
        """
        # normaliza nomes esperados (aliases)
        # Remove parênteses dos nomes de colunas para normalização (C(t) -> ct)
        cols = {c.lower().replace("(", "").replace(")", ""): c for c in df_norm.columns}
        well_col = cols.get("well", cols.get("well_id", cols.get("poco", cols.get("poÃ§o", ""))))
        sample_col = cols.get("samplename", cols.get("sample_name", cols.get("amostra", cols.get("sample", ""))))
        code_col = cols.get("codigo", cols.get("code", sample_col))
        target_col = cols.get("target", cols.get("target name", cols.get("target_name", "")))
        ct_col = cols.get("ct", cols.get("ct_value", cols.get("ct_media", cols.get("ct mean", ""))))
        res_col = cols.get("resultado", cols.get("resultado_final", cols.get("result", "resultado")))

        records = []
        for _, r in df_norm.iterrows():
            w = str(r.get(well_col, "")).strip()
            if not w:
                continue
            sample = str(r.get(sample_col, "")).strip()
            code = str(r.get(code_col, sample)).strip()
            alvo = str(r.get(target_col, "")).strip()
            ct_val = r.get(ct_col, None)
            try:
                ct_val = float(str(ct_val).replace(",", "."))
            except Exception:
                ct_val = None
            res_val = r.get(res_col, "")
            records.append(
                {
                    "Poco": w,
                    "Amostra": sample,
                    "Codigo": code,
                    f"Resultado_{alvo}": normalize_result(res_val),
                    f"CT_{alvo}": ct_val,
                }
            )
        if not records:
            return pd.DataFrame()
        df_flat = pd.DataFrame(records)
        # agrega por poço mantendo primeiro valor para amostra/código e combinando colunas
        def _agg(series):
            # pega primeiro não vazio
            for v in series:
                if pd.notna(v) and str(v).strip():
                    return v
            return ""

        group_cols = [c for c in df_flat.columns if c.startswith("Resultado_") or c.startswith("CT_")]
        df_grouped = (
            df_flat.groupby("Poco")
            .agg({**{"Amostra": _agg, "Codigo": _agg}, **{c: _agg for c in group_cols}})
            .reset_index()
        )
        return df_grouped

    @staticmethod
    def _infer_group_size(df: pd.DataFrame) -> int:
        sizes = []
        # Buscar coluna Poco ou POCO corretamente
        poco_col = df.get("Poco")
        if poco_col is None:
            poco_col = df.get("POCO")
        
        if poco_col is None:
            return 1
            
        for v in poco_col.fillna(""):
            if v:
                sizes.append(len(str(v).split("+")))
        if not sizes:
            return 1
        # pega o tamanho mais frequente
        return max(set(sizes), key=sizes.count)

    @staticmethod
    def _detect_control(sample: Optional[str], code: Optional[str], exam_cfg: Optional[Any] = None) -> Optional[str]:
        """
        Detecta se a amostra/código representa um controle.
        Primeiro tenta usar `exam_cfg.controles` (se fornecido), senão faz heurística por nomes comuns.
        Retorna 'CN' ou 'CP' ou None.
        """
        vals = []
        if sample:
            vals.append(str(sample).upper())
        if code:
            vals.append(str(code).upper())

        # se a config do exame fornece listas de controles, compare contra elas
        try:
            if exam_cfg and getattr(exam_cfg, "controles", None):
                cn_list = []
                for x in (exam_cfg.controles.get("cn") or []):
                    cn_list.extend([w.strip().upper() for w in str(x).split('+')])
                
                cp_list = []
                for x in (exam_cfg.controles.get("cp") or []):
                    cp_list.extend([w.strip().upper() for w in str(x).split('+')])
                
                for v in vals:
                    if v in cn_list:
                        return "CN"
                    if v in cp_list:
                        return "CP"
        except Exception:
            pass

        # Fallback heuristics
        for v in vals:
            if v in {"CN", "CONTROLE NEGATIVO", "C-", "NEGATIVO CONTROLE", "NEGATIVO", "CONTROLE N"}:
                return "CN"
            if v in {"CP", "CONTROLE POSITIVO", "C+", "POSITIVO CONTROLE", "POSITIVO", "CONTROLE P"}:
                return "CP"
        return None

    def _recompute_status(self, well: WellData) -> None:
        """Determina o status do poco baseado APENAS nos resultados textuais dos alvos."""
        if well.is_control:
            ctype = well.metadata.get("control_type", "")
            well.status = CONTROL_CN if ctype == "CN" else CONTROL_CP
            return

        has_pos = False
        has_inc = False
        has_nd = False

        # Analisar cada alvo (exceto RP)
        for alvo, tr in well.targets.items():
            if alvo.upper().startswith("RP"):
                continue

            # Analisar APENAS resultado textual
            raw = repair_mojibake_text(tr.result or "").strip()
            u = unicodedata.normalize("NFKD", raw)
            u = "".join(ch for ch in u if not unicodedata.combining(ch)).upper()

            # Indeterminado deve ser avaliado antes de DET,
            # pois "INDETERMINADO" contem o trecho "DET".
            if "INC" in u or "INDETERMIN" in u:
                has_inc = True
            elif "ND" in u or "NAO" in u or "NEGATIVO" in u or u.strip() == "":
                has_nd = True
            elif "DET" in u or "POS" in u or "DETECTA" in u:
                has_pos = True

        # Determinar status final
        if has_inc:
            well.status = INDETERMINADO
        elif has_pos:
            well.status = POSITIVE
        elif has_nd:
            well.status = NEGATIVE
        else:
            well.status = INVALID

    # utilidades
    def get_well(self, well_id: str) -> Optional[WellData]:
        return self.wells.get(well_id)

    def get_group(self, well_id: str) -> List[str]:
        """Retorna lista de poços no mesmo grupo (exceto o próprio well_id)"""
        w = self.get_well(well_id)
        if not w:
            return []
        # Suporta ambos os sistemas: novo (group_id) e legado (pair_group_id)
        if w.group_id:
            return [x for x in self.group_dict.get(w.group_id, []) if x != well_id]
        elif w.pair_group_id:
            return [x for x in self.pair_groups.get(w.pair_group_id, []) if x != well_id]
        return []
    
    def get_group_wells_including_self(self, well_id: str) -> List[str]:
        """Retorna todos os poços do grupo incluindo o próprio well_id"""
        w = self.get_well(well_id)
        if not w:
            return []
        if w.group_id:
            return self.group_dict.get(w.group_id, [])
        elif w.pair_group_id:
            return self.pair_groups.get(w.pair_group_id, [])
        return [well_id]

    def recompute_all(self) -> None:
        for w in self.wells.values():
            self._recompute_status(w)
    
    
    def _calcular_resultado_geral(self, well: WellData) -> str:
        """Calcula Resultado_geral delegando a domain.resultado_geral.calcular_resultado_geral."""
        from domain.resultado_geral import calcular_resultado_geral

        if not well.targets:
            return "Sem Alvos"

        # Determinar validade do RP (mesma lógica de antes: INV ou sem "VAL" = inválido)
        rp_valido = True
        for alvo, tr in well.targets.items():
            alvo_upper = alvo.upper()
            if alvo_upper.startswith("RP") or "RP_" in alvo_upper or "RP-" in alvo_upper:
                raw = repair_mojibake_text(tr.result or "").strip()
                result_upper = unicodedata.normalize("NFKD", raw)
                result_upper = "".join(ch for ch in result_upper if not unicodedata.combining(ch)).upper()
                if "INV" in result_upper or (result_upper and "VAL" not in result_upper):
                    if result_upper not in ("", "ND", "DET"):
                        rp_valido = False
                        break

        # Construir dict {nome_alvo: resultado} excluindo RP e GERAL
        alvos: dict[str, str] = {}
        for alvo, tr in well.targets.items():
            alvo_upper = alvo.upper()
            if alvo_upper.startswith("RP") or "RP_" in alvo_upper or "RP-" in alvo_upper:
                continue
            if "GERAL" in alvo_upper:
                continue
            raw = repair_mojibake_text(tr.result or "").strip()
            alvos[alvo] = raw

        return calcular_resultado_geral(rp_valido, alvos)

    def to_dataframe(self) -> pd.DataFrame:
        """
        Converte o PlateModel de volta para um DataFrame no formato df_final AGRUPADO.
        
        CRÃTICO: Para exames de 48/32/24 testes, cada linha representa um GRUPO de poços
        (ex: "A01+A02" ao invés de linhas separadas para A01 e A02).
        
        Retorna DataFrame com colunas: Poco, Amostra, Codigo, Resultado_<ALVO>, CT_<ALVO>...
        Usa 'Poco' (não 'Poço') para compatibilidade com sistema.
        """
        records = []
        processed_groups = set()
        
        # Correção #6 (Padronização): Helper para mapear resultados
        def _map_result(val: str, target: str) -> str:
            if not val:
                return ""
            normalized = normalize_result(str(val))
            normalized_token = str(normalized).strip().upper()
            is_indeterminate = normalized_token == "INC" or normalized_token.startswith("INDETERMIN")

            # Regra para RP (Controle Interno)
            if "RP" in target.upper():
                if normalized_token == "DET":
                    return ResultStatus.VALIDO
                if normalized_token == "ND":
                    return ResultStatus.INVALIDO
                return val  # Retorna original se nao mapeado

            # Regra para Alvos Normais
            if normalized_token == "DET":
                return ResultStatus.DETECTAVEL
            if normalized_token == "ND":
                return ResultStatus.NAO_DETECTAVEL
            if is_indeterminate:
                return ResultStatus.INDETERMINADO

            return val  # Retorna original se nao mapeado

        def _map_result_full(val: str, target: str) -> str:
            """Mapeia resultado para texto completo (Resultado_*)."""
            if not val:
                return ""
            normalized = normalize_result(str(val))
            normalized_token = str(normalized).strip().upper()
            is_indeterminate = normalized_token == "INC" or normalized_token.startswith("INDETERMIN")

            if "RP" in target.upper():
                if normalized_token == "DET":
                    return "Válido"
                if normalized_token == "ND":
                    return "Inválido"
                return val

            if normalized_token == "DET":
                return "Detectado"
            if normalized_token == "ND":
                return "Não Detectado"
            if is_indeterminate:
                return "Inconclusivo"
            return val

# Correcao #5: Ordenar grupos por COLUNA primeiro (A1+A2, B1+B2, ..., H1+H2, A3+A4, ...)
        def _sort_key_por_coluna(group_id: str) -> tuple:
            """
            Chave de ordenacao para grupos no padrao esperado:
            A1+A2, B1+B2, ..., H1+H2, A3+A4, B3+B4, ..., H3+H4, ...
            
            Retorna: (coluna_min, linha)
            """
            wells = group_id.split('+')
            if not wells:
                return (99, 'Z')
            
            # Pegar primeira coluna do grupo (ex: "01" de "A01+A02")
            first_well = wells[0]
            linha = first_well[0]  # 'A'
            try:
                coluna = int(first_well[1:])  # 1
            except (ValueError, IndexError):
                coluna = INVALID_WELL_COLUMN
            
            return (coluna, linha)
        
        # Ordenar grupos por coluna primeiro
        sorted_groups = sorted(self.group_dict.items(), key=lambda x: _sort_key_por_coluna(x[0]))
        
        # Processar grupos primeiro (para exames de 48/32/24 testes)
        for group_id, well_ids in sorted_groups:
            if group_id in processed_groups:
                continue
            
            # Pegar primeiro poço do grupo como representante
            if not well_ids:
                continue
            
            first_well_id = well_ids[0]
            first_well = self.wells.get(first_well_id)
            if not first_well:
                continue
            
            # Criar registro para o GRUPO
            # FASE 2.1: Normalizar group_id (A01+A02 â†’ A1+A2)
            normalized_group_id = "+".join([normalize_well_id(w) for w in group_id.split("+")])
            
            record = {
                "Poco": normalized_group_id,  # Normalizado
                "Amostra": first_well.sample_id or "",
                "Codigo": first_well.code or "",
            }
            
            # Adicionar resultados e CTs de cada alvo
            # Usar dados do primeiro poço (todos no grupo têm os mesmos resultados)
            for target_name, target_result in first_well.targets.items():
                target_clean = target_name.strip()
                
                # PULAR GERAL (Não gerar colunas CT_GERAL ou Res_GERAL)
                if target_clean.upper() == "GERAL":
                    continue

                # Mapear valor para padrão do sistema (Res_)
                mapped_val = _map_result(target_result.result, target_clean)
                record[f"Res_{target_clean}"] = mapped_val
                if self.include_resultado_cols:
                    record[f"Resultado_{target_clean}"] = _map_result_full(target_result.result, target_clean)

                if target_result.ct is not None:
                    record[f"CT_{target_clean}"] = target_result.ct
                else:
                    record[f"CT_{target_clean}"] = ""

                record[f"Manual_{target_clean}"] = bool(target_result.resultado_manual)
            
            if self.include_resultado_geral:
                # CORREÇÃO CRÃTICA: Calcular e exportar Resultado_geral
                resultado_geral = self._calcular_resultado_geral(first_well)
                record["Resultado_geral"] = resultado_geral
            
            records.append(record)
            processed_groups.add(group_id)
        
        # Processar poços individuais (não agrupados) - para exames de 96 testes
        for well_id, well in self.wells.items():
            # Pular se já foi processado como parte de um grupo
            if well.is_grouped and well.group_id in processed_groups:
                continue
            
            # Pular poços vazios
            if well.status == EMPTY and not well.sample_id:
                continue
            
            # Criar registro para poço individual
            # FASE 2.1: Normalizar well_id (A01 â†’ A1)
            normalized_well_id = normalize_well_id(well_id)
            
            record = {
                "Poco": normalized_well_id,  # Normalizado
                "Amostra": well.sample_id or "",
                "Codigo": well.code or "",
            }
            
            # Adicionar resultados e CTs
            for target_name, target_result in well.targets.items():
                target_clean = target_name.strip()
                
                # PULAR GERAL
                if target_clean.upper() == "GERAL":
                    continue

                # Mapear valor para padrão do sistema (Res_)
                mapped_val = _map_result(target_result.result, target_clean)
                record[f"Res_{target_clean}"] = mapped_val
                if self.include_resultado_cols:
                    record[f"Resultado_{target_clean}"] = _map_result_full(target_result.result, target_clean)

                if target_result.ct is not None:
                    record[f"CT_{target_clean}"] = target_result.ct
                else:
                    record[f"CT_{target_clean}"] = ""

                record[f"Manual_{target_clean}"] = bool(target_result.resultado_manual)
            
            if self.include_resultado_geral:
                # CORREÇÃO CRÃTICA: Calcular e exportar Resultado_geral
                resultado_geral = self._calcular_resultado_geral(well)
                record["Resultado_geral"] = resultado_geral
            
            records.append(record)
        
        if not records:
            return pd.DataFrame()
        
        return pd.DataFrame(records)
    
    def _determine_exam_type(self) -> None:
        """Determina o tipo de exame (96, 48, 32, 24 testes) baseado nos tamanhos de grupo"""
        if not self.group_dict:
            self.exam_type = "96"
            return
        
        group_sizes = {}
        for wells in self.group_dict.values():
            size = len(wells)
            group_sizes[size] = group_sizes.get(size, 0) + 1
        
        if not group_sizes:
            self.exam_type = "96"
            return
        
        # Encontrar tamanho mais comum corretamente
        most_common_size = max(group_sizes.items(), key=lambda x: x[1])[0]
        
        if most_common_size == 2:
            self.exam_type = "48"
        elif most_common_size == 3:
            self.exam_type = "32"
        elif most_common_size == 4:
            self.exam_type = "24"
        else:
            self.exam_type = "96"
    
    def _determine_group_frame_requirement(self) -> None:
        """Determina se é necessário criar frames de grupo com contorno"""
        self.requires_group_frames = self.exam_type in ["48", "32", "24"]


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------


def normalize_result(value: str) -> str:
    """Normaliza textos de resultado do CSV (ex: 'SC2 - 1', 'HMPV - 2')."""
    if not value:
        return ""

    raw = repair_mojibake_text(value).strip()
    txt = unicodedata.normalize("NFKD", raw)
    txt = "".join(ch for ch in txt if not unicodedata.combining(ch)).upper()

    # Formato especifico do CSV: "ALVO - NUMERO" (ex: "SC2 - 1", "HMPV - 2")
    if " - " in txt:
        parts = txt.split(" - ")
        if len(parts) >= 2:
            num = parts[-1].strip()
            if num == "1":
                return "Det"
            if num == "2":
                return "ND"
            return "Indeterminado"

    # Verificar termos mais especificos primeiro para evitar matches incorretos
    if any(k in txt for k in ["INC", "INDETERMIN", "3"]):
        return "Indeterminado"
    if any(k in txt for k in ["NAO DETECTADO", "NAO DETECTAVEL", "NEGATIVO"]):
        return "ND"
    if any(k in txt for k in ["DETECTADO", "DETECTAVEL", "POSITIVO", "REAGENTE", "1"]):
        return "Det"
    if any(k in txt for k in ["NAO", "NEGATIVO", "ND", "2"]):
        return "ND"

    return raw


def _controle_tem_alvo_detectavel(targets) -> bool:
    """True se algum alvo analitico (nao-RP) do controle for Detectavel/Positivo."""
    for target_name, target_result in (targets or {}).items():
        tu = str(target_name).upper()
        if tu.startswith("RP") or "RP_" in tu or "RP-" in tu:
            continue
        raw = repair_mojibake_text(str(getattr(target_result, "result", "") or "")).strip()
        ru = unicodedata.normalize("NFKD", raw)
        ru = "".join(ch for ch in ru if not unicodedata.combining(ch)).upper()
        if "INDETERMIN" in ru or "INC" in ru:
            continue
        if "NAO" in ru or ru == "ND" or "NEGATIVO" in ru:
            continue
        if "DET" in ru or "POS" in ru:
            return True
    return False


def controle_valido(targets, ctrl: str) -> bool:
    """Validade canonica de controle (espelha mapa_placa_exporter._validar_controle):
    CN valido = sem alvo detectavel (ND); CP valido = ao menos um alvo detectavel."""
    tem_detectavel = _controle_tem_alvo_detectavel(targets)
    if ctrl == "CN":
        return not tem_detectavel
    if ctrl == "CP":
        return tem_detectavel
    return True


def resolve_well_color(well: WellData) -> str:
    """
    Resolve cor do poco baseado nos resultados normalizados.

    Args:
        well: Estrutura com alvos e resultados.

    Returns:
        Cor em HEX.
    """
    if well.is_control:
        sample_upper = str(well.sample_id).upper()
        ctrl = "CN" if "CN" in sample_upper else ("CP" if "CP" in sample_upper else None)
        if ctrl is not None:
            # Controle invalido recebe marcacao de erro; valido recebe a cor trocada.
            if not controle_valido(well.targets, ctrl):
                return STATUS_COLORS[INVALID]
            return STATUS_COLORS[CONTROL_CN] if ctrl == "CN" else STATUS_COLORS[CONTROL_CP]

    if not well.targets:
        return STATUS_COLORS[EMPTY]

    has_positive = False
    has_indeterminado = False
    has_invalid = False

    for target_name, target_result in well.targets.items():
        target_upper = target_name.upper()
        if target_upper.startswith("RP") or "RP_" in target_upper or "RP-" in target_upper:
            continue

        raw_result = repair_mojibake_text(str(target_result.result or "")).strip()
        result_upper = unicodedata.normalize("NFKD", raw_result)
        result_upper = "".join(ch for ch in result_upper if not unicodedata.combining(ch)).upper()

        if "INV" in result_upper or "INVAL" in result_upper:
            has_invalid = True
        elif "INC" in result_upper or "INDETERMIN" in result_upper:
            has_indeterminado = True
        elif (
            "NAO" in result_upper
            or result_upper == "ND"
            or "NEGATIVO" in result_upper
            or result_upper.strip() == ""
        ):
            continue
        elif "DET" in result_upper or "POS" in result_upper or "DETECTA" in result_upper:
            has_positive = True

    if has_invalid:
        return STATUS_COLORS[INVALID]
    if has_indeterminado:
        return STATUS_COLORS[INDETERMINADO]
    if has_positive:
        return STATUS_COLORS[POSITIVE]
    return STATUS_COLORS[NEGATIVE]


class WellButton(ctk.CTkButton):
    def __init__(self, master, well_id: str, text: str, color: str, on_click=None, **kwargs):
        super().__init__(master, **kwargs)
        self.well_id = well_id
        self.on_click_callback = on_click
        self.group_size = 1
        self.group_position = 0
        self.is_group_highlight = False
        
        # Configuração do botão
        self.configure(
            width=60,
            height=50,
            fg_color=color,
            text_color="black",
            font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"),
            corner_radius=4,
            border_width=1,
            border_color="black",
            text=self._truncate(text, 10),
            command=self._on_click
        )
    
    def _truncate(self, text: str, max_length: int) -> str:
        """Trunca texto se muito longo."""
        if len(text) > max_length:
            return text[:max_length-2] + ".."
        return text
    
    def _on_click(self):
        """Manipula clique no botão."""
        if self.on_click_callback:
            self.on_click_callback(self.well_id)

    def update_appearance(self, text: str, color: str, is_selected: bool, 
                         is_group_highlight: bool = False, group_size: int = 1,
                         group_position: int = 0):
        """Atualiza a aparência do botão."""
        # Atualizar cor de fundo e texto
        self.configure(
            fg_color=color,
            text=self._truncate(text, 10)
        )
        
        # Atualizar propriedades do grupo
        self.group_size = group_size
        self.group_position = group_position
        self.is_group_highlight = is_group_highlight
        
        # Definir borda com prioridades: selecionado > destaque de grupo > normal
        if is_selected:
            self.configure(border_color="#FF0000", border_width=3)  # Vermelho para selecionado
        elif is_group_highlight:
            # Cor específica para o tamanho do grupo
            group_color = GROUP_COLORS.get(group_size, "#00AA00")
            self.configure(border_color=group_color, border_width=2)
        else:
            # Para poços em grupos, manter borda normal
            if group_size > 1:
                self.configure(border_color="#888888", border_width=1)
            else:
                self.configure(border_color="#888888", border_width=2)


class GroupFrame(ctk.CTkFrame):
    """Frame que agrupa múltiplos poços para contorná-los juntos."""
    
    def __init__(self, master, group_size: int, border_color: str, corner_radius: int = 15):
        # Configurar o frame com borda de 3px
        super().__init__(
            master,
            fg_color="transparent",
            border_width=3,
            border_color=border_color,
            corner_radius=corner_radius
        )
        self.group_size = group_size
        
        # Ajustar o layout baseado no tamanho do grupo
        if group_size == 2:
            # Para pares, organizar horizontalmente (1x2)
            self.grid_columnconfigure(0, weight=1)
            self.grid_columnconfigure(1, weight=1)
            self.grid_rowconfigure(0, weight=1)
        elif group_size == 3:
            # Para trios, organizar em linha (1x3)
            for i in range(3):
                self.grid_columnconfigure(i, weight=1)
            self.grid_rowconfigure(0, weight=1)
        elif group_size == 4:
            # Para quartetos, organizar em 2x2
            for i in range(2):
                self.grid_columnconfigure(i, weight=1)
                self.grid_rowconfigure(i, weight=1)
    
    def get_position_in_group(self, position: int) -> tuple:
        """Retorna (row, col) dentro do GroupFrame baseado na posição do poço no grupo."""
        if self.group_size == 2:
            # Par: posição 0 -> (0, 0), posição 1 -> (0, 1)
            return (0, position)
        elif self.group_size == 3:
            # Trio: posição 0 -> (0, 0), posição 1 -> (0, 1), posição 2 -> (0, 2)
            return (0, position)
        elif self.group_size == 4:
            # Quarteto 2x2: posição 0 -> (0, 0), posição 1 -> (0, 1), posição 2 -> (1, 0), posição 3 -> (1, 1)
            return (position // 2, position % 2)
        return (0, 0)


class PlateView(ctk.CTkFrame):
    def __init__(self, master, plate_model: PlateModel, meta: Dict[str, str], on_save_callback=None):
        super().__init__(master)
        self.plate_model = plate_model
        self.meta = meta or {}
        # DEPRECATED: Mantido para compatibilidade, mas não usado
        self.on_save_callback = on_save_callback
        
        # NOVO (P2 Fix): Usar Event Bus para desacoplar comunicação
        from services.core.event_bus import EventBus, SystemEvents
        self.event_bus = EventBus
        
        # CORRECAO CRITICA: Inicializar atributos AQUI no __init__ (estavam dentro de _get_well_color!)
        self.selected_well_id: Optional[str] = None
        self.current_target: Optional[str] = None
        self.highlight_group: List[str] = []
        self.group_wells_highlight: List[str] = []  # Lista para destacar grupo
        self.group_size_highlight: int = 1
        self.well_widgets: Dict[str, WellButton] = {}
        self.group_frames: Dict[str, GroupFrame] = {}

        self._build_ui()
        
        # FASE 1 - CORREÇÃO #2: Renderizar cores após UI estar pronta (evita loop infinito)
        self.after(50, self.render_plate)
    
    def _get_well_color(self, well: WellData) -> str:
        """
        Retorna cor de fundo baseada no status/resultados do well.

        CORRECAO CRITICA: GERAL nao e um alvo, e Resultado_geral (coluna calculada).
        Esta funcao analisa APENAS os alvos reais (ADV, HMPV, RSV, etc.) para determinar a cor.
        """
        from utils.logger import registrar_log
        registrar_log("DEBUG_COLOR", f"Well {well.well_id}: is_control={well.is_control}, targets={list(well.targets.keys()) if well.targets else 'None'}", "DEBUG")
        cor_final = resolve_well_color(well)
        registrar_log("DEBUG_COLOR", f"Well {well.well_id}: cor={cor_final}", "DEBUG")
        return cor_final

    def _build_ui(self):
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Cabeçalho compacto - 2 linhas no topo
        header = ctk.CTkFrame(self)
        header.grid(row=0, column=0, columnspan=2, sticky="ew", padx=(5,500), pady=2)
        
        # Linha 1: Data, Extracao, Exame
        linha1 = f"Data: {self.meta.get('data', '')} | Extracao: {self.meta.get('extracao', self.meta.get('arquivo', ''))} | Exame: {self.meta.get('exame', '')}"
        ctk.CTkLabel(header, text=linha1, font=("Segoe UI", 15), anchor="w").grid(row=0, column=0, sticky="ew", padx=5)
        
        # Linha 2: Usuário, Tamanho do bloco
        linha2 = f"Usuario: {self.meta.get('usuario', '')} | Tamanho bloco: {self.plate_model.group_size} (Total amostras: {self._calc_total_samples()})"
        ctk.CTkLabel(header, text=linha2, font=("Segoe UI", 15, "bold"), anchor="w").grid(row=1, column=0, sticky="ew", padx=5)
        
        header.grid_columnconfigure(0, weight=0)

        # Container da Placa
        plate_container = ctk.CTkFrame(self)
        plate_container.grid(row=1, column=0, padx=(1,10), pady=(1, 1), sticky="nsew")
        plate_container.grid_rowconfigure(0, weight=1)
        plate_container.grid_columnconfigure(0, weight=1)

        self.plate_frame = ctk.CTkFrame(plate_container)
        self.plate_frame.grid(row=0, column=0, sticky="nsew")

        # Títulos colunas (1-12)
        font_labels = ctk.CTkFont(family="Segoe UI", size=11, weight="bold")
        ctk.CTkLabel(self.plate_frame, text="", width=30, height=30).grid(row=0, column=0, padx=1, pady=1)
        for j, col in enumerate(COL_LABELS, start=1):
            label = ctk.CTkLabel(
                self.plate_frame, 
                text=col, 
                font=font_labels,
                width=90,
                height=30
            )
            label.grid(row=0, column=j, padx=1, pady=1)
        
        # Rótulos de linha (A-H)
        for i, row_lbl in enumerate(ROW_LABELS, start=1):
            label = ctk.CTkLabel(
                self.plate_frame,
                text=row_lbl,
                font=font_labels,
                width=30,
                height=40
            )
            label.grid(row=i, column=0, padx=1, pady=1)

        # Criar frames de grupo se necessário
        if self.plate_model.requires_group_frames:
            self._create_group_frames()
        
        # Criar botões de poços
        self._create_well_buttons()

    def _create_group_frames(self):
        """Cria frames com bordas coloridas para agrupar poços."""
        for group_id, wells in self.plate_model.group_dict.items():
            if not wells:
                continue
            
            # Determinar tamanho do grupo
            group_size = len(wells)
            if group_size not in GROUP_COLORS:
                continue
            
            color = GROUP_COLORS[group_size]
            
            # Calcular posição mínima e máxima do grupo
            rows = [ROW_LABELS.index(w[0]) for w in wells]
            # Remover zeros Ã  esquerda da coluna (A01 -> 1, A12 -> 12)
            cols = [COL_LABELS.index(str(int(w[1:]))) for w in wells]
            min_row = min(rows)
            max_row = max(rows)
            min_col = min(cols)
            max_col = max(cols)
            
            # Determinar corner_radius baseado no tamanho do grupo
            corner_radius_map = {2: 12, 3: 15, 4: 18}
            corner_radius = corner_radius_map.get(group_size, 15)
            
            # Criar o GroupFrame
            group_frame = GroupFrame(
                self.plate_frame,
                group_size=group_size,
                border_color=color,
                corner_radius=corner_radius
            )
            
            # Posicionar no grid (+1 para compensar os labels)
            group_frame.grid(
                row=min_row + 1,
                column=min_col + 1,
                rowspan=max_row - min_row + 1,
                columnspan=max_col - min_col + 1,
                padx=1,
                pady=1,
                sticky="nsew"
            )
            
            self.group_frames[group_id] = group_frame

    def _create_well_buttons(self):
        """Cria todos os botões de poços, usando frames de grupo quando necessário."""
        for i, row_lbl in enumerate(ROW_LABELS):
            for j, col_lbl in enumerate(COL_LABELS):
                # Criar well_id com zero Ã  esquerda (A01, A02, etc)
                well_id = f"{row_lbl}{int(col_lbl):02d}"
                well = self.plate_model.get_well(well_id)
                
                # Preparar texto do botão
                text = ""
                if well:
                    text = well.code or well.sample_id or ""
                    if text and well.is_control:
                        ct_type = well.metadata.get("control_type", "")
                        if ct_type:
                            text = f"{ct_type}:{text}"
                
                # Determinar cor baseada nos resultados (CORRECAO: colorização dinâmica)
                color = self._get_well_color(well) if well else STATUS_COLORS[EMPTY]
                
                # Determinar frame pai e posição no grid
                if well and well.is_grouped and self.plate_model.requires_group_frames:
                    parent_frame = self.group_frames.get(well.group_id)
                    if parent_frame:
                        # Calcular posição dentro do grupo
                        grid_row, grid_col = parent_frame.get_position_in_group(well.group_position)
                        grid_kwargs = {"row": grid_row, "column": grid_col}
                    else:
                        # Fallback se não encontrar o frame
                        parent_frame = self.plate_frame
                        grid_kwargs = {"row": i + 1, "column": j + 1}
                else:
                    parent_frame = self.plate_frame
                    grid_kwargs = {"row": i + 1, "column": j + 1}
                
                # Criar botão com espaçamento entre eles
                btn = WellButton(parent_frame, well_id, text, color, on_click=self.on_well_click)
                btn.grid(padx=1, pady=1, sticky="nsew", **grid_kwargs)
                self.well_widgets[well_id] = btn

        # Botão Salvar e Voltar na janela principal
        btn_container = ctk.CTkFrame(self, fg_color="transparent")
        btn_container.grid(row=2, column=0, pady=10)
        
        ctk.CTkLabel(btn_container, text="Clique na amostra que deseja editar.", font=("Segoe UI", 14, "italic")).pack(side="left", padx=20)
        
        ctk.CTkButton(
            btn_container,
            text="Salvar Alteracoes e Voltar",
            font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
            height=40,
            fg_color="#27AE60",
            hover_color="#229954",
            command=self._salvar_e_voltar,
        ).pack(side="left", padx=20)

    def _ensure_detail_popup(self):
        if hasattr(self, '_detail_popup') and self._detail_popup and self._detail_popup.winfo_exists():
            return
            
        self._detail_popup = ctk.CTkToplevel(self)
        self._detail_popup.title("Edição de Amostra")
        self._detail_popup.geometry("520x720")
        self._detail_popup.transient(self.winfo_toplevel())
        
        self.detail_frame = ctk.CTkFrame(self._detail_popup)
        self.detail_frame.pack(fill="both", expand=True, padx=5, pady=5)
        self.detail_frame.grid_propagate(False)

        # Configuração das linhas e colunas do painel de detalhes
        self.detail_frame.grid_columnconfigure(0, weight=1)
        self.detail_frame.grid_columnconfigure(1, weight=1)
        for i in range(9):
            self.detail_frame.grid_rowconfigure(i, weight=0)
        self.detail_frame.grid_rowconfigure(6, weight=1) # TreeView expande

        # Seção de Informações
        title_font = ctk.CTkFont(family="Segoe UI", size=16, weight="bold")
        ctk.CTkLabel(self.detail_frame, text="Poco selecionado:", font=title_font).grid(
            row=0, column=0, columnspan=2, pady=(2, 1), padx=15, sticky="w"
        )

        f_label = ctk.CTkFont(family="Segoe UI", size=12, weight="bold")
        f_val = ctk.CTkFont(family="Segoe UI", size=12)

        # Campos info: pady=1px, padx=(15, 5)px e (5, 15)px
        ctk.CTkLabel(self.detail_frame, text="Poco:", font=f_label).grid(row=1, column=0, padx=(15, 5), pady=1, sticky="e")
        self.lbl_well = ctk.CTkLabel(self.detail_frame, text="-", font=f_val)
        self.lbl_well.grid(row=1, column=1, padx=(5, 15), pady=1, sticky="w")

        ctk.CTkLabel(self.detail_frame, text="Amostra:", font=f_label).grid(row=2, column=0, padx=(15, 5), pady=1, sticky="e")
        sample_frame = ctk.CTkFrame(self.detail_frame, fg_color="transparent")
        sample_frame.grid(row=2, column=1, padx=(5, 15), pady=3, sticky="ew")
        sample_frame.grid_columnconfigure(0, weight=1)
        self.entry_sample = ctk.CTkEntry(sample_frame, font=f_val, height=35)
        self.entry_sample.grid(row=0, column=0, sticky="ew", padx=(0, 5))
        ctk.CTkButton(sample_frame, text="OK", width=40, height=35, font=f_val, command=self.apply_sample_change).grid(
            row=0, column=1
        )

        ctk.CTkLabel(self.detail_frame, text="Codigo:", font=f_label).grid(row=3, column=0, padx=(15, 5), pady=1, sticky="e")
        code_frame = ctk.CTkFrame(self.detail_frame, fg_color="transparent")
        code_frame.grid(row=3, column=1, padx=(5, 15), pady=3, sticky="ew")
        code_frame.grid_columnconfigure(0, weight=1)
        self.entry_code = ctk.CTkEntry(code_frame, font=f_val, height=35) # Entry código: height=35px
        self.entry_code.grid(row=0, column=0, sticky="ew", padx=(0, 5))
        ctk.CTkButton(code_frame, text="OK", width=40, height=35, font=f_val, command=self.apply_code_change).grid(
            row=0, column=1
        )

        ctk.CTkLabel(self.detail_frame, text="Pocos agrupados:", font=f_label).grid(
            row=4, column=0, padx=(15, 5), pady=1, sticky="e"
        )
        self.lbl_group = ctk.CTkLabel(self.detail_frame, text="-", font=f_val)
        self.lbl_group.grid(row=4, column=1, padx=(5, 15), pady=1, sticky="w")

        ctk.CTkLabel(self.detail_frame, text="Resultados:", font=f_label).grid(
            row=5, column=0, columnspan=2, padx=15, pady=(1, 1), sticky="w"
        )

        # TreeView (Tabela de Resultados)
        tree_frame = ctk.CTkFrame(self.detail_frame)
        tree_frame.grid(row=6, column=0, columnspan=2, padx=15, pady=(0, 5), sticky="nsew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        style = ttk.Style()
        style.theme_use("default")
        style.configure(
            "Treeview",
            foreground="black",
            background="white",
            fieldbackground="white",
            font=("Segoe UI", 15), # Fonte: 15px
            rowheight=42, # Rowheight: 42px
        )
        style.configure("Treeview.Heading", foreground="black", background="#f0f0f0", font=("Segoe UI", 15, "bold"))
        # Altura adequada: 5 linhas para visualização compacta e evitar clipping
        self.tree = ttk.Treeview(tree_frame, columns=("alvo", "resultado", "ct"), show="headings", selectmode="browse", height=5)
        self.tree.heading("alvo", text="Alvo")
        self.tree.heading("resultado", text="Resultado")
        self.tree.heading("ct", text="CT")
        # Colunas: 65px / 95px / 70px
        self.tree.column("alvo", width=65, anchor="w")
        self.tree.column("resultado", width=95, anchor="center")
        self.tree.column("ct", width=70, anchor="center")
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        self.tree.bind("<ButtonRelease-1>", self.on_tree_select)

        # Edit Frame
        edit_frame = ctk.CTkFrame(self.detail_frame)
        edit_frame.grid(row=7, column=0, columnspan=2, padx=15, pady=(2, 2), sticky="ew")
        edit_frame.grid_columnconfigure(1, weight=1)
        edit_frame.grid_columnconfigure(3, weight=1)
        
        # Campo de Alvo (editável)
        ctk.CTkLabel(edit_frame, text="Alvo:", font=f_label).grid(row=0, column=0, padx=5, pady=2, sticky="e")
        self.entry_target = ctk.CTkEntry(edit_frame, width=120, font=f_val, height=35) # Entries: height=35px
        self.entry_target.grid(row=0, column=1, padx=5, pady=2, sticky="ew")
        
        # Campo de Resultado
        ctk.CTkLabel(edit_frame, text="Resultado:", font=f_label).grid(row=1, column=0, padx=5, pady=2, sticky="e")
        self.entry_res = ctk.CTkEntry(edit_frame, width=90, font=f_val, height=35) # Entries: height=35px
        self.entry_res.grid(row=1, column=1, padx=5, pady=2, sticky="ew")
        
        # Campo de CT
        ctk.CTkLabel(edit_frame, text="CT:", font=f_label).grid(row=1, column=2, padx=(10, 5), pady=2, sticky="e")
        self.entry_ct = ctk.CTkEntry(edit_frame, width=90, font=f_val, height=35) # Entries: height=35px
        self.entry_ct.grid(row=1, column=3, padx=5, pady=2, sticky="ew")
        
        # Botão Aplicar
        ctk.CTkButton(
            edit_frame, text="Aplicar", font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), height=38, command=self.apply_target_changes
        ).grid(row=2, column=0, columnspan=4, pady=(5, 2))

        # Botao Fechar
        ctk.CTkButton(
            self.detail_frame,
            text="Fechar Janela",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
            height=40,
            fg_color="#E74C3C",
            hover_color="#C0392B",
            command=self._detail_popup.withdraw,
        ).grid(row=8, column=0, columnspan=2, padx=15, pady=(0, 4))

    def _calc_total_samples(self) -> int:
        # total de grupos = total_wells / group_size
        if not self.plate_model.wells:
            return 0
        return math.ceil(len(self.plate_model.wells) / max(1, self.plate_model.group_size))

    def _status_color(self, status: str) -> str:
        return STATUS_COLORS.get(status, "#ffffff")

    # ------------------ interação ------------------ #
    def on_well_click(self, well_id: str) -> None:
        self.selected_well_id = well_id
        
        # Obter informações do grupo
        well = self.plate_model.get_well(well_id)
        if well and well.is_grouped:
            group_wells = self.plate_model.get_group_wells_including_self(well_id)
            self.group_wells_highlight = list(set(group_wells))  # Converter set para list
            self.group_size_highlight = well.group_size
        else:
            self.group_wells_highlight = []  # Lista vazia em vez de set
            self.group_size_highlight = 0
        
        self.render_plate()
        if well:
            self._ensure_detail_popup()
            self._detail_popup.deiconify()
            self._detail_popup.lift()
            self._fill_details(well)

    def render_plate(self):
        for well_id, btn in self.well_widgets.items():
            well = self.plate_model.get_well(well_id)
            text = ""
            if well:
                text = well.code or well.sample_id or ""
                if text and well.is_control:
                    ctype = well.metadata.get("control_type", "")
                    if ctype:
                        text = f"{ctype}:{text}"
            
            # CORRECAO CRITICA: usar _get_well_color em vez de _status_color
            color = self._get_well_color(well) if well else STATUS_COLORS[EMPTY]
            is_selected = (well_id == self.selected_well_id)
            is_group_highlight = (well_id in self.group_wells_highlight)
            
            # Passar informações de grupo para o botão
            group_size = 0
            group_position = 0
            if well and well.is_grouped:
                group_size = well.group_size
                group_position = well.group_position
            
            btn.update_appearance(text, color, is_selected, is_group_highlight, group_size, group_position)

    def _fill_details(self, well: WellData):
        self.lbl_well.configure(text=well.well_id)
        self.entry_sample.delete(0, tk.END)
        if well.sample_id:
            self.entry_sample.insert(0, well.sample_id)
        self.entry_code.delete(0, tk.END)
        if well.code:
            self.entry_code.insert(0, well.code)
        self.lbl_group.configure(text=", ".join(well.paired_wells) if well.paired_wells else "-")

        self.tree.delete(*self.tree.get_children())
        # ordenar alvos: primeiro não-RP, depois RPs
        def _sort_key(item):
            name = item[0]
            is_rp = name.upper().startswith("RP")
            return (1 if is_rp else 0, name)

        for alvo, tr in sorted(well.targets.items(), key=_sort_key):
            # CORREÇÃO 2026-02-09: Ignorar "GERAL" na interface de edição (coluna calculada)
            if alvo.upper() == "GERAL":
                continue

            # Usar vírgula como separador decimal
            ct_txt = "" if tr.ct is None else f"{tr.ct:.3f}".replace(".", ",")
            res_display = f"{tr.result} (M)" if tr.resultado_manual else tr.result
            item_id = self.tree.insert("", "end", values=(alvo, res_display, ct_txt))
            self.tree.item(item_id, tags=(alvo,))
        self.entry_target.delete(0, tk.END)
        self.entry_res.delete(0, tk.END)
        self.entry_ct.delete(0, tk.END)
        self.current_target = None

        # Selecionar automaticamente o primeiro item se houver
        children = self.tree.get_children()
        if children:
            self.tree.selection_set(children[0])
            self.tree.focus(children[0])
            # Chamar on_tree_select para carregar os dados automaticamente
            self.on_tree_select(None)

    def on_tree_select(self, event):
        sel = self.tree.selection()
        if not sel:
            self.current_target = None
            return
        vals = self.tree.item(sel[0], "values")
        if not vals:
            return
        self.current_target = vals[0]
        
        # Popular campo de alvo
        self.entry_target.delete(0, tk.END)
        self.entry_target.insert(0, vals[0])
        
        # Popular campo de resultado (limpar o sufixo ' (M)' se houver para evitar duplicações/erros)
        self.entry_res.delete(0, tk.END)
        res_val = str(vals[1]).strip()
        if res_val.endswith(" (M)"):
            res_val = res_val[:-4]
        elif res_val.endswith("(M)"):
            res_val = res_val[:-3]
        self.entry_res.insert(0, res_val)
        
        # Popular campo de CT (já está com vírgula)
        self.entry_ct.delete(0, tk.END)
        self.entry_ct.insert(0, vals[2])

    def apply_sample_change(self):
        if not self.selected_well_id:
            return
        well = self.plate_model.get_well(self.selected_well_id)
        if not well:
            return
        new_sample = self.entry_sample.get().strip()
        well.sample_id = new_sample or None
        ctrl = PlateModel._detect_control(well.sample_id, well.code, self.plate_model.exam_cfg)
        if ctrl:
            well.is_control = True
            well.metadata["control_type"] = ctrl
        else:
            well.is_control = False
            well.metadata.pop("control_type", None)
        # se amostra mudou, propaga para wells do mesmo grupo
        group_wells = self.plate_model.get_group_wells_including_self(well.well_id)
        for wid in group_wells:
            if wid == well.well_id:
                continue
            w2 = self.plate_model.get_well(wid)
            if w2:
                w2.sample_id = well.sample_id
                w2.code = well.code
                if ctrl:
                    w2.is_control = True
                    w2.metadata["control_type"] = ctrl
                else:
                    w2.is_control = False
                    w2.metadata.pop("control_type", None)
                self.plate_model._recompute_status(w2)

        self.plate_model._recompute_status(well)
        self.render_plate()

    def apply_code_change(self):
        if not self.selected_well_id:
            return
        well = self.plate_model.get_well(self.selected_well_id)
        if not well:
            return
        new_code = self.entry_code.get().strip()
        well.code = new_code or None
        ctrl = PlateModel._detect_control(well.sample_id, well.code, self.plate_model.exam_cfg)
        if ctrl:
            well.is_control = True
            well.metadata["control_type"] = ctrl
        else:
            well.is_control = False
            well.metadata.pop("control_type", None)
        # se code mudou, propaga para wells do mesmo grupo (mesma amostra)
        group_wells = self.plate_model.get_group_wells_including_self(well.well_id)
        for wid in group_wells:
            if wid == well.well_id:
                continue
            w2 = self.plate_model.get_well(wid)
            if w2:
                w2.code = well.code
                w2.sample_id = well.sample_id
                if ctrl:
                    w2.is_control = True
                    w2.metadata["control_type"] = ctrl
                else:
                    w2.is_control = False
                    w2.metadata.pop("control_type", None)
                self.plate_model._recompute_status(w2)

        self.plate_model._recompute_status(well)
        self.render_plate()

    def apply_target_changes(self):
        """
        Aplica alteracoes de alvo/resultado/CT ao poco selecionado.
        
        IMPORTANTE: Esta funcao atualiza APENAS o PlateModel em memoria.
        Para sincronizar com a aba de analise, use o botao "Salvar Alteracoes e Voltar"
        que chama _salvar_e_voltar() â†’ on_save_callback() â†’ _on_mapa_salvo() da janela principal.
        """
        if not self.selected_well_id or not self.current_target:
            return
        well = self.plate_model.get_well(self.selected_well_id)
        if not well:
            return
        
        # Obter valores originais para comparação
        original_target = well.targets.get(self.current_target)
        original_res = original_target.result if original_target else ""
        original_ct = original_target.ct if original_target else None
        original_manual = original_target.resultado_manual if original_target else False

        # Obter novos valores
        new_target_name = self.entry_target.get().strip()
        new_res = normalize_result(self.entry_res.get())
        ct_text = self.entry_ct.get().strip()
        new_ct = None
        if ct_text:
            try:
                # Aceitar tanto vírgula quanto ponto como separador decimal
                new_ct = float(ct_text.replace(",", "."))
            except Exception:
                new_ct = None

        original_res_normalized = normalize_result(original_res)
        auto_recalculated = False

        # Se apenas o CT foi alterado (resultado não mudou), reavaliar o resultado baseado no CT
        if new_res == original_res_normalized and new_ct != original_ct:
            # CORREÇÃO 2026-02-09: Usar mesmas regras de classificar_ct (logic_engine.py)
            # Regras Biomanguinhos VR1e2: <8=ND, 8-35=Det, 35-40=Inc, >40=ND, vazio=ND
            if new_ct is None or new_ct == '':
                new_res = "ND"
            elif new_ct < 8:
                new_res = "ND"
            elif new_ct < 35:
                new_res = "Det"
            elif new_ct <= 40:
                new_res = "Inc"
            else:  # > 40
                new_res = "ND"
            auto_recalculated = True

        # Determinar flag de override manual:
        # - True quando usuario alterou resultado explicitamente (nao por recalculo de CT)
        # - Preserva flag anterior quando nada mudou
        result_changed = new_res != original_res_normalized
        if result_changed:
            is_manual = not auto_recalculated
        else:
            is_manual = original_manual

        # Se o nome do alvo mudou, remover o antigo
        if new_target_name and new_target_name != self.current_target:
            if self.current_target in well.targets:
                del well.targets[self.current_target]

        # Atualizar o alvo (novo ou existente)
        target_key = new_target_name if new_target_name else self.current_target
        well.targets[target_key] = TargetResult(new_res, new_ct, resultado_manual=is_manual)

        # Reanalisar este poço pelas regras
        self.plate_model._recompute_status(well)

        # Propagar mudanças para todos os poços do grupo
        if well.is_grouped:
            group_wells = self.plate_model.get_group_wells_including_self(self.selected_well_id)
            for well_id in group_wells:
                if well_id == self.selected_well_id:
                    continue  # Já atualizamos o poço atual
                w2 = self.plate_model.get_well(well_id)
                if w2:
                    # Obter valores originais do poço do grupo
                    original_target_w2 = w2.targets.get(self.current_target)
                    original_res_w2 = original_target_w2.result if original_target_w2 else ""
                    original_ct_w2 = original_target_w2.ct if original_target_w2 else None

                    # Aplicar a mesma lógica de reanálise se apenas CT mudou
                    new_res_w2 = new_res
                    auto_recalculated_w2 = False
                    if normalize_result(original_res_w2) == normalize_result(original_res) and new_ct != original_ct_w2:
                        # CORREÇÃO 2026-02-09: Mesmas regras de classificar_ct
                        if new_ct is None or new_ct == '':
                            new_res_w2 = "ND"
                        elif new_ct < 8:
                            new_res_w2 = "ND"
                        elif new_ct < 35:
                            new_res_w2 = "Det"
                        elif new_ct <= 40:
                            new_res_w2 = "Inc"
                        else:  # > 40
                            new_res_w2 = "ND"
                        auto_recalculated_w2 = True

                    result_changed_w2 = new_res_w2 != normalize_result(original_res_w2)
                    original_manual_w2 = original_target_w2.resultado_manual if original_target_w2 else False
                    is_manual_w2 = (not auto_recalculated_w2) if result_changed_w2 else original_manual_w2

                    # Se o nome do alvo mudou, remover o antigo
                    if new_target_name and new_target_name != self.current_target:
                        if self.current_target in w2.targets:
                            del w2.targets[self.current_target]

                    # Atualizar o alvo no poço do grupo
                    w2.targets[target_key] = TargetResult(new_res_w2, new_ct, resultado_manual=is_manual_w2)

                    # Reanalisar este poço também
                    self.plate_model._recompute_status(w2)
        
        # Atualizar o current_target para o novo nome
        self.current_target = target_key
        
        # Atualizar interface
        self._fill_details(well)
        self.render_plate()

    def _on_save_clicked(self):
            """
            Salva as alterações (Sincroniza) e FECHA a janela para liberar o menu.
            """
            # 1. Recupera o AppState da janela mãe (se injetado) ou via master
            app_state = getattr(self.master, "app_state", None)
            
            if app_state and app_state.resultados_analise is not None:
                try:
                    # Se você implementou o sync_to_dataframe (da análise anterior):
                    # df_atualizado = self.plate_model.sync_to_dataframe(app_state.resultados_analise)
                    # app_state.resultados_analise = df_atualizado
                    pass # Substitua pelo código de sync real
                except Exception as e:
                    print(f"Erro ao sincronizar: {e}")

            # 2. COMANDO CRÃTICO: Fechar a janela
            # Isso encerra o wait_window() no MenuHandler e destrava o sistema.
            self.master.destroy()
    
    def _salvar_e_voltar(self):
        """
        Salva alterações e notifica parent (não destrói mais a janela).
        
        NOVO COMPORTAMENTO (janela única com abas):
        - Se parent é JanelaAnaliseCompleta: apenas notifica via callback
        - Se parent é PlateWindow (legado): destrói Toplevel normalmente
        """
        try:
            # Recomputar todos os status antes de salvar
            self.plate_model.recompute_all()
            
            # Executar callback se fornecido
            # P2 FIX: Publicar evento em vez de chamar callback diretamente
            from services.core.event_bus import EventBus, SystemEvents
            EventBus.publish(SystemEvents.PLATE_SAVED, {
                "plate_model": self.plate_model,
                "metadata": self.meta
            })
            
            # Manter callback para compatibilidade com código legacy
            if self.on_save_callback:
                self.on_save_callback(self.plate_model)
            
            # CRÃTICO: Verificar tipo do parent para decidir comportamento
            toplevel = self.winfo_toplevel()
            
            # Se parent é CTkTabview ou Frame: NÃO destruir (sistema de abas)
            # Se parent é PlateWindow (CTkToplevel): destruir (sistema legado)
            if isinstance(toplevel, ctk.CTkToplevel) and type(toplevel).__name__ == "PlateWindow":
                # Sistema legado: destruir Toplevel
                self._destruir_toplevel_seguro(toplevel)
            else:
                # Sistema de abas: parent controla navegação, não fazemos nada
                from utils.logger import registrar_log
                registrar_log("PlateView", "Alterações salvas (sistema de abas)", "INFO")
                
        except Exception as e:
            from utils.logger import registrar_log
            registrar_log("PlateView", f"Erro ao salvar: {e}", "ERROR")
            from tkinter import messagebox
            messagebox.showerror(
                "Erro",
                f"Falha ao salvar alteracoes:\n{str(e)}",
                parent=self
            )
    
    def _destruir_toplevel_seguro(self, toplevel):
        """
        Destrói Toplevel de forma segura (apenas para sistema legado).
        """
        try:
            # Ocultar imediatamente
            toplevel.withdraw()
            
            # Agendar destruição após delay
            def destruir():
                try:
                    if toplevel.winfo_exists():
                        toplevel.destroy()
                except Exception:
                    pass
            
            toplevel.after(200, destruir)
            
        except Exception as e:
            from utils.logger import registrar_log
            registrar_log("PlateView", f"Erro ao destruir janela: {e}", "ERROR")


class PlateWindow(ctk.CTkToplevel):
    def __init__(self, root, plate_model: PlateModel, meta: Dict[str, str], on_save_callback=None):
        # Importar AfterManagerMixin para gerenciar callbacks
        from utils.after_mixin import AfterManagerMixin
        
        # Adicionar suporte ao AfterManagerMixin via composição
        self._after_ids = set()
        
        super().__init__(master=root)
        self.title("Visualizacao da Placa")
        
        # Vincular ao parent (mantém janela acima do parent)
        self.transient(root)
        
        # Definir tamanho inicial (90% da tela para maximizar área da placa)
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        window_width = int(screen_width * 0.9)
        window_height = int(screen_height * 0.9)
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Proteção contra TclError ao fechar
        self._is_closing = False
        self.protocol("WM_DELETE_WINDOW", self._on_close_window)
        
        # Estrutura Geral da Janela: Padding externo: padx=10px, pady=2px
        view = PlateView(self, plate_model, meta, on_save_callback=on_save_callback)
        view.pack(fill="both", expand=True, padx=10, pady=2)
    
    def dispose(self):
        """Cancela todos os callbacks agendados."""
        for aid in self._after_ids:
            try:
                self.after_cancel(aid)
            except Exception:
                pass
        self._after_ids.clear()
    
    def schedule(self, delay_ms: int, callback, *args, **kwargs):
        """Agendar callback e registrar para cancelamento posterior."""
        aid = self.after(delay_ms, callback, *args, **kwargs)
        self._after_ids.add(aid)
        return aid
    
    def _on_close_window(self):
        """Fecha a janela com segurança."""
        if not self._is_closing:
            self._is_closing = True
            try:
                # Cancelar callbacks pendentes
                self.dispose()
                
                # Ocultar janela imediatamente
                self.withdraw()
                
                # Destruir após delay para callbacks internos do CustomTkinter
                def destruir_seguro():
                    try:
                        if self.winfo_exists():
                            self.destroy()
                    except Exception:
                        pass
                
                self.after(200, destruir_seguro)
            except Exception:
                pass


# ---------------------------------------------------------------------------
# API pública
# ---------------------------------------------------------------------------


def abrir_placa_ctk(df_final: pd.DataFrame, meta_extra: Optional[Dict[str, Any]] = None, group_size: Optional[int] = None, parent=None, on_save_callback=None):
    """
    Abre a janela CTk para visualização/edição da placa usando df_final em memória.
    meta_extra pode conter data, extracao/arquivo, exame, usuario.
    on_save_callback: função a ser chamada ao salvar alterações (recebe PlateModel).
    
    IMPORTANTE: parent deve sempre ser passado para evitar criação de segundo root CTk.
    """
    try:
        # CRÃTICO: Validar parent para prevenir criação de segundo root CTk
        # Criar ctk.CTk() quando já existe mainloop ativo causa travamentos
        if parent is None:
            # Em testes, permitimos parent=None (PlateWindow costuma ser mockado)
            if os.environ.get("PYTEST_CURRENT_TEST"):
                parent = None
            else:
                raise RuntimeError(
                    "abrir_placa_ctk requer um parent CTk/CTkToplevel válido.\n"
                    "Passar parent=None criaria um segundo root, causando travamento da aplicação.\n"
                    "Solução: Sempre passe a janela principal como parent."
                )
        
        print(f"DEBUG abrir_placa_ctk: DataFrame shape={df_final.shape if df_final is not None else 'None'}")
        
        if df_final is None or df_final.empty:
            print("DEBUG abrir_placa_ctk: DataFrame vazio ou None")
            return
        
        meta = meta_extra or {}
        # garantir chaves esperadas
        meta.setdefault("data", meta.get("data_placa", ""))
        meta.setdefault("extracao", meta.get("arquivo_corrida", meta.get("extracao", "")))
        meta.setdefault("exame", meta.get("exame", ""))
        meta.setdefault("usuario", meta.get("usuario", ""))
        
        print(f"DEBUG abrir_placa_ctk: meta={meta}, group_size={group_size}")
        
        # Passa exame para PlateModel.from_df para carregação do registry
        exame = meta.get("exame", "")
        
        print(f"DEBUG abrir_placa_ctk: Criando PlateModel.from_df...")
        plate_model = PlateModel.from_df(df_final, group_size=group_size, exame=exame)
        
        print(f"DEBUG abrir_placa_ctk: PlateModel criado, wells={len(plate_model.wells)}")
        
        print(f"DEBUG abrir_placa_ctk: Criando PlateWindow...")
        # parent já foi validado acima, não pode ser None
        win = PlateWindow(parent, plate_model, meta, on_save_callback=on_save_callback)
        win.focus_force()
        
        print(f"DEBUG abrir_placa_ctk: PlateWindow criada com sucesso")
        return win
    except Exception as e:
        print(f"DEBUG abrir_placa_ctk: ERRO - {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        raise


# Compatibilidade legada: funções vazias (Excel removido nesta fase)
def construir_well_results(*args, **kwargs):
    raise NotImplementedError("Função legada não suportada nesta versão.")


def exportar_placa_excel(wells: List[WellResult], output_path: str) -> None:
    """Exporta mapa da placa para Excel a partir da lista de wells."""
    from pathlib import Path

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    def _well_to_position(well_id: str) -> tuple[int, int]:
        normalized = normalize_well_id(well_id)
        row_label = normalized[0]
        col_num = int(normalized[1:])
        row_idx = ROW_LABELS.index(row_label) + 2
        col_idx = col_num + 1
        return row_idx, col_idx

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "MapaPlaca"

    ws["A1"] = "Well"
    ws["A1"].font = Font(bold=True)
    for col in range(1, 13):
        cell = ws.cell(row=1, column=col + 1, value=str(col))
        cell.font = Font(bold=True)
    for row, label in enumerate(ROW_LABELS, start=2):
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(bold=True)

    for item in wells:
        if not item.well:
            continue
        try:
            row_idx, col_idx = _well_to_position(item.well)
        except Exception:
            continue
        lines = [str(item.sample_code or "").strip(), str(item.result or "").strip()]
        if item.ct_target is not None:
            lines.append(f"CT={item.ct_target:.2f}")
        if item.ct_rp is not None:
            lines.append(f"RP={item.ct_rp:.2f}")
        value = "\n".join([line for line in lines if line])
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    col_names = "ABCDEFGHIJKLM"
    ws.column_dimensions["A"].width = 6
    for idx in range(2, 14):
        ws.column_dimensions[col_names[idx - 1]].width = 16
    for row in range(2, 10):
        ws.row_dimensions[row].height = 48

    wb.save(out)


def mostrar_placa_gui(*args, **kwargs):
    # wrapper simples para seguir nomes antigos
    return abrir_placa_ctk(*args, **kwargs)


# Configuração inicial do tema
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")


# ---------------------------------------------------------------------------
# Compatibilidade: carregar PlateModel a partir de historico_analises.csv
# ---------------------------------------------------------------------------


@classmethod
def _plate_model_from_historico_csv(cls, csv_path: str, sep: str = ";", exame: str | None = None, arquivo_corrida: str | None = None):
    import pandas as pd
    from pathlib import Path

    with CSVFileLock(csv_path):
        df = pd.read_csv(csv_path, sep=sep)

    # filtros opcionais
    col_map = {str(c).strip().lower(): c for c in df.columns}
    col_exame = col_map.get('exame')
    col_arquivo = col_map.get('arquivo_corrida')

    if exame and col_exame:
        df = df[df[col_exame].astype(str).str.strip().str.casefold() == str(exame).strip().casefold()]
    if arquivo_corrida and col_arquivo:
        base = Path(str(arquivo_corrida)).name
        df = df[df[col_arquivo].astype(str).str.strip().str.casefold() == base.strip().casefold()]

    model = cls()
    upper_cols = [str(c).upper() for c in df.columns]
    model.include_resultado_cols = any(
        c.startswith("RESULTADO_") and c != "RESULTADO_GERAL" for c in upper_cols
    )
    model.include_resultado_geral = "RESULTADO_GERAL" in upper_cols

    # identifica colunas basicas
    poco_col = col_map.get('poco') or col_map.get('poço') or col_map.get('well')
    amostra_col = col_map.get('amostra') or col_map.get('sample')
    codigo_col = col_map.get('codigo') or col_map.get('código') or col_map.get('code')

    # precomputar colunas de resultado/ct
    result_cols = []
    ct_cols = []
    for c in df.columns:
        cu = str(c).strip()
        cu_upper = cu.upper()
        if cu_upper.endswith(' - R'):
            result_cols.append((c, cu_upper[:-3].strip()))
        elif cu_upper.startswith('RESULTADO_'):
            result_cols.append((c, cu_upper.replace('RESULTADO_', '').strip()))
        elif cu_upper.startswith('RES_'):
            result_cols.append((c, cu_upper.replace('RES_', '').strip()))
        elif cu_upper.endswith(' - CT'):
            ct_cols.append((c, cu_upper[:-4].strip()))
        elif cu_upper.startswith('CT_'):
            ct_cols.append((c, cu_upper.replace('CT_', '').strip()))
        elif cu_upper.startswith('RP') and 'CT' not in cu_upper and 'RESULTADO' not in cu_upper and 'RES_' not in cu_upper:
            ct_cols.append((c, cu_upper.strip()))

    def _parse_ct(val):
        if val is None:
            return None
        try:
            if pd.isna(val):
                return None
        except Exception:
            pass
        s = str(val).strip()
        if not s:
            return None
        try:
            return float(s.replace(',', '.'))
        except Exception:
            return None

    def _valid_well_id(well_id: str) -> bool:
        if not well_id or len(well_id) < 2:
            return False
        row = well_id[0]
        col = well_id[1:]
        if row not in ROW_LABELS:
            return False
        if not col.isdigit():
            return False
        col_num = int(col)
        return 1 <= col_num <= 12

    def _detect_control(sample: str | None, code: str | None):
        token = (sample or code or '').strip().upper()
        if not token:
            return None
        if token == 'CN' or token.startswith('CN') or 'CONTROL NEG' in token or token.startswith('NEG'):
            return 'CN'
        if token == 'CP' or token.startswith('CP') or 'CONTROL POS' in token or token.startswith('POS'):
            return 'CP'
        return None

    def _classify_result(value: str | None):
        if value is None:
            return None
        token = classify_result_text(value)
        if token == "DET":
            return POSITIVE
        if token == "ND":
            return NEGATIVE
        if token == "INC":
            return INDETERMINADO
        if token == "INV":
            return INVALID
        return None

    for _, row in df.iterrows():
        poco_raw = str(row.get(poco_col, '')).strip() if poco_col else ''
        if not poco_raw:
            continue
        tokens = [p.strip() for p in poco_raw.split('+') if p.strip()]
        if not tokens:
            continue

        sample_val = str(row.get(amostra_col, '')).strip() if amostra_col else ''
        code_val = str(row.get(codigo_col, '')).strip() if codigo_col else ''

        for token in tokens:
            well_id = normalize_well_id(token)
            if not _valid_well_id(well_id):
                continue

            row_label = well_id[0]
            col_label = well_id[1:]
            well = model.wells.get(well_id)
            if well is None:
                well = WellData(row_label=row_label, col_label=col_label, well_id=well_id)
                model.wells[well_id] = well

            if sample_val and not well.sample_id:
                well.sample_id = sample_val
            if code_val and not well.code:
                well.code = code_val

            # preencher targets
            for col, target in result_cols:
                if target.upper() == 'GERAL':
                    continue
                val = row.get(col, None)
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    continue
                tr = well.targets.get(target, TargetResult())
                tr.result = str(val).strip()
                well.targets[target] = tr

            for col, target in ct_cols:
                if target.upper() == 'GERAL':
                    continue
                ct_val = _parse_ct(row.get(col, None))
                if ct_val is None:
                    continue
                tr = well.targets.get(target, TargetResult())
                tr.ct = ct_val
                well.targets[target] = tr

            well.metadata.update({
                'exame': row.get(col_exame) if col_exame else None,
                'arquivo_corrida': row.get(col_arquivo) if col_arquivo else None,
            })

    # definir status/controle
    for well in model.wells.values():
        ctrl = _detect_control(well.sample_id, well.code)
        if ctrl == 'CN':
            well.is_control = True
            well.status = CONTROL_CN
            continue
        if ctrl == 'CP':
            well.is_control = True
            well.status = CONTROL_CP
            continue

        well.is_control = False
        has_pos = False
        has_inc = False
        has_neg = False
        for target, tr in well.targets.items():
            if str(target).upper() == 'GERAL':
                continue
            cls_val = _classify_result(tr.result)
            if cls_val == POSITIVE:
                has_pos = True
            elif cls_val == INDETERMINADO:
                has_inc = True
            elif cls_val == NEGATIVE:
                has_neg = True
        if has_pos:
            well.status = POSITIVE
        elif has_inc:
            well.status = INDETERMINADO
        elif has_neg:
            well.status = NEGATIVE
        else:
            well.status = INVALID

    return model


# registrar no PlateModel
PlateModel.from_historico_csv = _plate_model_from_historico_csv

