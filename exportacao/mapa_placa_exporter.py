# -*- coding: utf-8 -*-
"""
Exporter do Mapa de Placa definitivo (.xlsx P&B paisagem A4).

Le um pandas.DataFrame de analise + ExamConfig e gera o mapa visual
seguindo as especificacoes do laboratorio:

  - Cabecalho com nome do exame + validacao da placa (PLACA OK / RECOMECAR)
  - 8 linhas de blocos × N blocos-por-linha (N = 12 / pocos_por_amostra)
  - Cada bloco: codigo amostra (16pt bold) + grid 3 col fonte 7pt + badge resultado
  - Rodape: tabela horizontal compacta de controles + campos de assinatura

Toda regra de classificacao e distribuicao de alvos vive em domain.mapa_placa_layout.
"""
from __future__ import annotations

import math
import os
import unicodedata
from datetime import datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins, PrintOptions

from domain.mapa_placa_layout import (
    CLASSIF_DETECTAVEL,
    CLASSIF_INDETERMINADO,
    CLASSIF_INVALIDO,
    CLASSIF_NAO_DETECTAVEL,
    AlvoCelula,
    BlocoAmostra,
    ControleQuadro,
    MapaPlaca,
    classificar_amostra,
    distribuir_alvos_grid,
    formatar_ct,
    montar_mapa,
)

# ----------------------------------------------------------------------------
# Estilos
# ----------------------------------------------------------------------------

_BLACK = "FF000000"
_WHITE = "FFFFFFFF"
_GRAY = "FF999999"
_LIGHT_GRAY = "FFE0E0E0"

_THIN_BORDER = Side(style="thin", color=_BLACK)
_BLOCK_BORDER = Border(
    left=_THIN_BORDER, right=_THIN_BORDER, top=_THIN_BORDER, bottom=_THIN_BORDER
)

_FILL_BLACK = PatternFill("solid", fgColor=_BLACK)
_FILL_WHITE = PatternFill("solid", fgColor=_WHITE)
_FILL_GRAY = PatternFill("solid", fgColor=_GRAY)
_FILL_RESULT_INVALID = PatternFill("solid", fgColor="FFF0F0F0")
_FILL_HEADER_LIGHT = PatternFill("solid", fgColor=_LIGHT_GRAY)

_GROUP_BORDER_COLORS = {
    2: "FF3498DB",
    3: "FFE74C3C",
    4: "FFF39C12",
}

_STATUS_FILLS = {
    "NEGATIVE": PatternFill("solid", fgColor="FFD4F4D4"),
    "POSITIVE": PatternFill("solid", fgColor="FFFFB3B3"),
    "INDETERMINADO": PatternFill("solid", fgColor="FFFFE0B2"),
    "INVALID": PatternFill("solid", fgColor="FFF0F0F0"),
    "CONTROL_CN": PatternFill("solid", fgColor="FFB3D9FF"),
    "CONTROL_CP": PatternFill("solid", fgColor="FFB3D9FF"),
    "EMPTY": _FILL_WHITE,
}

_RESULTADO_GERAL_COLS = {"resultado_geral", "resultadogeral", "resultado geral"}


def _font(size: int, bold: bool = False, white: bool = False) -> Font:
    return Font(
        name="Arial",
        size=size,
        bold=bold,
        color=(_WHITE if white else _BLACK),
    )


# Cada bloco ocupa 3 colunas Excel × 5 linhas Excel
# linhas dentro do bloco:
#   0: codigo da amostra (merge 3 cols)
#   1-3: grid 3x3 de alvos (linha do grid em 1 linha excel)
#   4: badge resultado (merge 3 cols)
BLOCK_COLS = 3
BLOCK_ROWS = 5

PHYSICAL_GRID_HEADER_ROW = 4
PHYSICAL_GRID_FIRST_ROW = 5
PHYSICAL_GRID_LABEL_COL = 1
PHYSICAL_GRID_FIRST_COL = 2
PHYSICAL_WELL_ROW_STEP = 6
PHYSICAL_WELL_COL_STEP = 2
PHYSICAL_TOTAL_COLS = PHYSICAL_GRID_FIRST_COL + ((12 - 1) * PHYSICAL_WELL_COL_STEP)
PHYSICAL_BLOCK_ROWS = 5


# ----------------------------------------------------------------------------
# Adapter df_analise -> blocos
# ----------------------------------------------------------------------------


def _extrair_alvos_e_rps(colunas: Sequence[str]) -> Tuple[List[str], List[str]]:
    """Separa colunas CT_* em alvos comuns e RPs (ou CIs)."""
    alvos: List[str] = []
    rps: List[str] = []
    for c in colunas:
        if not isinstance(c, str) or not c.startswith("CT_"):
            continue
        nome = c[3:]
        if nome.startswith("RP") or nome.startswith("CI"):
            rps.append(nome)
        else:
            alvos.append(nome)
    return alvos, rps


def _ct_da_linha(linha: pd.Series, alvo: str) -> Any:
    col = f"CT_{alvo}"
    if col not in linha.index:
        return None
    val = linha[col]
    if val is None:
        return None
    if isinstance(val, float) and math.isnan(val):
        return None
    if isinstance(val, str) and not val.strip():
        return None
    return val


def _res_da_linha(linha: pd.Series, alvo: str) -> str:
    for col in (f"Res_{alvo}", f"Resultado_{alvo}"):
        if col in linha.index:
            return str(linha[col] or "")
    return ""


def _normalizar_token_resultado(valor: object) -> str:
    texto = str(valor or "").strip().casefold()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return texto


def _rp_valido_da_linha(linha: pd.Series, rps: Sequence[str]) -> bool:
    """RP valido se TODOS os Res_RP_* / Res_CI_* contiverem token valido."""
    if not rps:
        return True
    aceitos = {"detectavel", "detectado", "det", "valido", "valida", "v"}
    rejeitados = {"invalido", "invalida", "inv"}
    for rp in rps:
        col_res = next(
            (col for col in (f"Res_{rp}", f"Resultado_{rp}") if col in linha.index),
            None,
        )
        if col_res is None:
            return False
        val = _normalizar_token_resultado(linha[col_res])
        if not val:
            return False
        if any(token in val for token in rejeitados):
            return False
        if val not in aceitos and not any(token in val for token in aceitos):
            return False
    return True


def _resultado_geral_da_linha(linha: pd.Series) -> str:
    for col in linha.index:
        if str(col).strip().casefold() in _RESULTADO_GERAL_COLS:
            return str(linha[col] or "")
        if _normalizar_nome_coluna(col) == "resultadogeral":
            return str(linha[col] or "")
    return ""


def _classificar_por_resultado_geral(resultado: str) -> Optional[Tuple[str, Tuple[str, ...]]]:
    norm = _normalizar_token_resultado(resultado)
    if not norm:
        return None
    if "inval" in norm:
        return CLASSIF_INVALIDO, ()
    if "indeterm" in norm or "inconclus" in norm:
        return CLASSIF_INDETERMINADO, ()
    if "nao" in norm and "detect" in norm:
        return CLASSIF_NAO_DETECTAVEL, ()
    if "detect" in norm:
        detectaveis: Tuple[str, ...] = ()
        lower = str(resultado or "").lower()
        if "para" in lower:
            after = lower.split("para", 1)[1]
            detectaveis = tuple(
                parte.strip(" .;,\t").upper()
                for parte in after.replace(" e ", ",").split(",")
                if parte.strip(" .;,\t")
            )
        return CLASSIF_DETECTAVEL, detectaveis
    return None


def _linha_eh_controle(amostra: str) -> Optional[str]:
    """Retorna 'CN' ou 'CP' se a amostra parece controle; None caso contrario."""
    s = str(amostra or "").strip().upper()
    if not s:
        return None
    if s.startswith("CN") or "CONTROLE NEG" in s or "CONTROLE-NEG" in s:
        return "CN"
    if s.startswith("CP") or "CONTROLE POS" in s or "CONTROLE-POS" in s:
        return "CP"
    return None


def _validar_controle(nome: str, ct: Any, alvo: str) -> bool:
    """Regra simples de validacao de controle (placeholder).

    CN: deve ser ND (CT vazio)
    CP: deve ter CT detectavel (qualquer valor numerico valido)
    """
    is_empty = ct is None or (isinstance(ct, str) and not ct.strip())
    try:
        is_nan = isinstance(ct, float) and math.isnan(ct)
    except Exception:
        is_nan = False
    if nome == "CN":
        return is_empty or is_nan
    if nome == "CP":
        return not (is_empty or is_nan)
    return True


def construir_mapa_de_dataframe(
    df_analise: pd.DataFrame,
    nome_exame: str,
    nome_placa: str,
    pocos_por_amostra: int,
    ordem_alvos: Optional[Sequence[str]] = None,
) -> MapaPlaca:
    """Adapta df_analise -> MapaPlaca pronto para o exporter."""
    if df_analise is None or df_analise.empty:
        return montar_mapa(nome_exame, nome_placa, pocos_por_amostra, [], [])

    colunas = list(df_analise.columns)
    alvos_descobertos, rps_descobertos = _extrair_alvos_e_rps(colunas)

    if ordem_alvos:
        alvos_ordenados: List[str] = []
        for a in ordem_alvos:
            chave = a.replace(" ", "_")
            if chave in alvos_descobertos and chave not in alvos_ordenados:
                alvos_ordenados.append(chave)
        for a in alvos_descobertos:
            if a not in alvos_ordenados:
                alvos_ordenados.append(a)
    else:
        alvos_ordenados = alvos_descobertos

    blocos_lineares: List[BlocoAmostra] = []
    controles_acumulados: List[ControleQuadro] = []

    coluna_amostra = "Amostra" if "Amostra" in df_analise.columns else "Sample"
    coluna_poco = next(
        (
            c
            for c in df_analise.columns
            if c in ("Poço(s)", "PoÃ§o(s)", "Poco(s)", "Poco")
        ),
        None,
    )

    for _, linha in df_analise.iterrows():
        codigo = str(linha.get(coluna_amostra, "") or "").strip()
        if not codigo:
            continue

        poco_label = ""
        if coluna_poco and coluna_poco in linha.index:
            poco_label = str(linha.get(coluna_poco, "") or "")

        ctrl = _linha_eh_controle(codigo)
        if ctrl is not None:
            for alvo in alvos_ordenados:
                ct_val = _ct_da_linha(linha, alvo)
                ct_fmt = formatar_ct(ct_val)
                valido = _validar_controle(ctrl, ct_val, alvo)
                # so adiciona controle nao vazio (CN com ct vazio tambem entra)
                controles_acumulados.append(
                    ControleQuadro(
                        nome=ctrl,
                        alvo=alvo.replace("_", " "),
                        ct_formatado=ct_fmt,
                        valido=valido,
                    )
                )
        alvos_celulas: List[AlvoCelula] = []
        resultados: Dict[str, str] = {}
        for alvo in alvos_ordenados:
            ct = _ct_da_linha(linha, alvo)
            ct_fmt = formatar_ct(ct)
            alvos_celulas.append(
                AlvoCelula(nome=alvo.replace("_", " "), ct_formatado=ct_fmt)
            )
            resultados[alvo.replace("_", " ")] = _res_da_linha(linha, alvo)
        for rp in rps_descobertos:
            ct = _ct_da_linha(linha, rp)
            ct_fmt = formatar_ct(ct)
            alvos_celulas.append(
                AlvoCelula(nome=rp.replace("_", " "), ct_formatado=ct_fmt)
            )

        resultado_geral = _classificar_por_resultado_geral(_resultado_geral_da_linha(linha))
        if resultado_geral is not None and ctrl is None:
            classificacao, detectaveis = resultado_geral
        else:
            rp_valido = _rp_valido_da_linha(linha, rps_descobertos)
            classificacao, detectaveis = classificar_amostra(resultados, rp_valido)

        alvos_distribuidos = distribuir_alvos_grid(alvos_celulas)

        blocos_lineares.append(
            BlocoAmostra(
                codigo=codigo,
                alvos=alvos_distribuidos,
                classificacao=classificacao,
                detectaveis=detectaveis,
                poco_label=poco_label,
            )
        )

    return montar_mapa(
        nome_exame=nome_exame,
        nome_placa=nome_placa,
        pocos_por_amostra=pocos_por_amostra,
        blocos_lineares=blocos_lineares,
        controles=controles_acumulados,
    )


# ----------------------------------------------------------------------------
# Renderizador openpyxl
# ----------------------------------------------------------------------------


def _aplicar_borda(ws, r1: int, c1: int, r2: int, c2: int, border: Border = _BLOCK_BORDER) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = border


def _normalizar_nome_coluna(nome: object) -> str:
    texto = str(nome or "").strip().casefold()
    texto = (
        texto.replace("ã§", "c")
        .replace("ã‡", "c")
        .replace("Ã§", "c")
        .replace("Ã‡", "c")
        .replace("ï¿½", "c")
    )
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return "".join(ch for ch in texto if ch.isalnum())


def _normalizar_dataframe_para_plate_model(df_analise: pd.DataFrame) -> pd.DataFrame:
    """Normaliza aliases de colunas para o mesmo contrato consumido pela aba visual."""
    if df_analise is None:
        return pd.DataFrame()

    df = df_analise.copy()
    renames: Dict[object, str] = {}
    for coluna in df.columns:
        chave = _normalizar_nome_coluna(coluna)
        if chave in {"poco", "pocos", "well", "wellid"}:
            renames[coluna] = "Poco"
        elif chave in {"codigo", "code"}:
            renames[coluna] = "Codigo"
        elif chave in {"amostra", "sample", "samplename", "samplename"}:
            renames[coluna] = "Amostra"

    if renames:
        df = df.rename(columns=renames)
    return df


def _well_excel_row(row_label: str) -> int:
    row_labels = ["A", "B", "C", "D", "E", "F", "G", "H"]
    return PHYSICAL_GRID_FIRST_ROW + row_labels.index(row_label) * PHYSICAL_WELL_ROW_STEP


def _well_excel_col(col_label: str) -> int:
    return PHYSICAL_GRID_FIRST_COL + (int(col_label) - 1) * PHYSICAL_WELL_COL_STEP


def _border_group(size: int) -> Border:
    color = _GROUP_BORDER_COLORS.get(size, _BLACK)
    side = Side(style="medium", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _texto_well(well: Any) -> str:
    if well is None:
        return ""
    texto = str(getattr(well, "code", "") or getattr(well, "sample_id", "") or "").strip()
    if texto and getattr(well, "is_control", False):
        control_type = str(getattr(well, "metadata", {}).get("control_type", "") or "").strip()
        if control_type:
            return f"{control_type}:{texto}"
    return texto


def _texto_status_well(well: Any) -> str:
    status = str(getattr(well, "status", "") if well is not None else "").upper()
    if status == "POSITIVE":
        return "DETECTAVEL"
    if status == "NEGATIVE":
        return "NAO DETECTAVEL"
    if status == "INDETERMINADO":
        return "INDETERMINADO"
    if status == "INVALID":
        return "INVALIDO"
    return ""


def _codigo_numerico(texto: object) -> str:
    return "".join(ch for ch in str(texto or "") if ch.isdigit())


def _nome_alvo_excel(nome: str) -> str:
    texto = str(nome or "").strip().replace("_", "")
    texto = texto.replace(" ", "")
    return texto.upper() if texto.upper().startswith("RP") else texto


def _texto_alvo_ct(alvo: AlvoCelula) -> str:
    if not alvo.nome:
        return ""
    nome = _nome_alvo_excel(alvo.nome)
    return f"{nome} - {alvo.ct_formatado}" if alvo.ct_formatado else f"{nome} -"


def _fill_resultado(classificacao: str) -> PatternFill:
    if classificacao == CLASSIF_DETECTAVEL:
        return _FILL_BLACK
    if classificacao == CLASSIF_NAO_DETECTAVEL:
        return _FILL_WHITE
    if classificacao == CLASSIF_INDETERMINADO:
        return _FILL_GRAY
    if classificacao == CLASSIF_INVALIDO:
        return _FILL_GRAY
    return _FILL_WHITE


def _indexar_blocos_por_codigo(mapa: MapaPlaca) -> Dict[str, BlocoAmostra]:
    blocos: Dict[str, BlocoAmostra] = {}
    for linha in mapa.blocos:
        for bloco in linha:
            if bloco is None:
                continue
            for chave in {str(bloco.codigo or "").strip(), _codigo_numerico(bloco.codigo)}:
                if chave:
                    blocos.setdefault(chave, bloco)
    return blocos


def _bloco_para_well(well: Any, blocos_por_codigo: Dict[str, BlocoAmostra]) -> Optional[BlocoAmostra]:
    if well is None:
        return None
    candidatos = [
        getattr(well, "code", ""),
        getattr(well, "sample_id", ""),
        _texto_well(well),
    ]
    for candidato in candidatos:
        texto = str(candidato or "").strip()
        if texto in blocos_por_codigo:
            return blocos_por_codigo[texto]
        digitos = _codigo_numerico(texto)
        if digitos in blocos_por_codigo:
            return blocos_por_codigo[digitos]
    return None


def _merge_if_needed(ws, row1: int, col1: int, row2: int, col2: int) -> None:
    if row1 != row2 or col1 != col2:
        ws.merge_cells(start_row=row1, start_column=col1, end_row=row2, end_column=col2)


def _clear_block(ws, row_ini: int, col_ini: int, row_fim: int, col_fim: int) -> None:
    for row in range(row_ini, row_fim + 1):
        for col in range(col_ini, col_fim + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = _FILL_WHITE
            cell.border = Border()


def _escrever_bloco_grade_fisica(
    ws,
    row_ini: int,
    col_ini: int,
    col_fim: int,
    bloco: Optional[BlocoAmostra],
    border: Border,
) -> None:
    row_fim = row_ini + PHYSICAL_BLOCK_ROWS - 1
    _clear_block(ws, row_ini, col_ini, row_fim, col_fim)
    _aplicar_borda(ws, row_ini, col_ini, row_fim, col_fim, border)

    if bloco is None:
        return

    _merge_if_needed(ws, row_ini, col_ini, row_ini, col_fim)
    codigo = str(bloco.codigo or "").strip()
    cell_codigo = ws.cell(row=row_ini, column=col_ini, value=codigo)
    cell_codigo.font = _font(10, bold=True)
    cell_codigo.alignment = Alignment(horizontal="center", vertical="center")

    alvos = list(bloco.alvos)
    while len(alvos) < 9:
        alvos.append(AlvoCelula(nome="", ct_formatado=""))

    target_cols = [
        col
        for col in range(col_ini, col_fim + 1)
        if (col - PHYSICAL_GRID_FIRST_COL) % PHYSICAL_WELL_COL_STEP == 0
    ][:3]
    if not target_cols:
        target_cols = [col_ini]

    for linha_idx in range(3):
        row = row_ini + 1 + linha_idx
        textos = [_texto_alvo_ct(alvos[linha_idx * 3 + col_idx]) for col_idx in range(3)]
        if len(target_cols) >= 3:
            for col_idx, col in enumerate(target_cols[:3]):
                cell = ws.cell(row=row, column=col, value=textos[col_idx])
                cell.font = _font(7)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            _merge_if_needed(ws, row, col_ini, row, col_fim)
            cell = ws.cell(row=row, column=col_ini, value="    ".join(t for t in textos if t))
            cell.font = _font(7)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    row_resultado = row_ini + 4
    _merge_if_needed(ws, row_resultado, col_ini, row_resultado, col_fim)
    
    raw_res = str(bloco.texto_resultado or "").upper().replace("DETECTÁVEL PARA ", "").replace("DETECTAVEL PARA ", "")
    if bloco.classificacao == CLASSIF_DETECTAVEL:
        alvos_limpos = [
            a.strip() for a in raw_res.replace(" E ", ",").split(",")
            if a.strip() and not a.strip().startswith("RP") and not a.strip().startswith("CI")
        ]
        raw_res = ", ".join(alvos_limpos)
        
    cell_resultado = ws.cell(row=row_resultado, column=col_ini, value=raw_res)
    cell_resultado.font = _font(
        8,
        bold=True,
        white=(bloco.classificacao == CLASSIF_DETECTAVEL),
    )
    cell_resultado.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    if bloco.classificacao == CLASSIF_INDETERMINADO:
        cell_resultado.fill = _FILL_HEADER_LIGHT
    else:
        cell_resultado.fill = _fill_resultado(bloco.classificacao)
        
    if bloco.classificacao == CLASSIF_INVALIDO:
        for r_iter in range(row_ini, row_fim + 1):
            for c_iter in range(col_ini, col_fim + 1):
                ws.cell(row=r_iter, column=c_iter).fill = _FILL_GRAY


def _escrever_bloco(
    ws,
    bloco: Optional[BlocoAmostra],
    row_ini: int,
    col_ini: int,
) -> None:
    """Escreve um bloco no Excel comecando em (row_ini, col_ini). Ocupa BLOCK_ROWS x BLOCK_COLS."""
    r_codigo = row_ini
    r_grid_inicio = row_ini + 1
    r_resultado = row_ini + 4
    c_fim = col_ini + BLOCK_COLS - 1

    # bordas externas do bloco
    _aplicar_borda(ws, row_ini, col_ini, row_ini + BLOCK_ROWS - 1, c_fim)

    if bloco is None:
        # bloco vazio (apenas borda)
        ws.merge_cells(
            start_row=row_ini, start_column=col_ini, end_row=row_ini, end_column=c_fim
        )
        return

    # Linha 1: codigo da amostra (merge 3 cols)
    ws.merge_cells(
        start_row=r_codigo, start_column=col_ini, end_row=r_codigo, end_column=c_fim
    )
    cell_cod = ws.cell(row=r_codigo, column=col_ini, value=bloco.codigo)
    cell_cod.font = _font(11, bold=True)
    cell_cod.alignment = Alignment(horizontal="center", vertical="center")

    # Linhas 2-4: grid 3x3 de alvos
    alvos = list(bloco.alvos)
    while len(alvos) < 9:
        alvos.append(AlvoCelula(nome="", ct_formatado=""))
    for i in range(min(9, len(alvos))):
        r_offset = i // BLOCK_COLS
        c_offset = i % BLOCK_COLS
        r = r_grid_inicio + r_offset
        c = col_ini + c_offset
        alvo = alvos[i]
        if alvo.nome:
            texto = (
                f"{alvo.nome} – {alvo.ct_formatado}"
                if alvo.ct_formatado
                else f"{alvo.nome} – "
            )
        else:
            texto = ""
        cell = ws.cell(row=r, column=c, value=texto)
        cell.font = _font(7)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Linha 5: badge resultado (merge 3 cols)
    ws.merge_cells(
        start_row=r_resultado,
        start_column=col_ini,
        end_row=r_resultado,
        end_column=c_fim,
    )
    
    raw_res = str(bloco.texto_resultado or "").upper().replace("DETECTÁVEL PARA ", "").replace("DETECTAVEL PARA ", "")
    if bloco.classificacao == CLASSIF_DETECTAVEL:
        alvos_limpos = [
            a.strip() for a in raw_res.replace(" E ", ",").split(",")
            if a.strip() and not a.strip().startswith("RP") and not a.strip().startswith("CI")
        ]
        raw_res = ", ".join(alvos_limpos)
        
    cell_res = ws.cell(row=r_resultado, column=col_ini, value=raw_res)
    cell_res.alignment = Alignment(horizontal="center", vertical="center")
    cell_res.font = _font(8, bold=True, white=(bloco.classificacao == CLASSIF_DETECTAVEL))
    
    if bloco.classificacao == CLASSIF_DETECTAVEL:
        cell_res.fill = _FILL_BLACK
    elif bloco.classificacao == CLASSIF_INDETERMINADO:
        cell_res.fill = _FILL_HEADER_LIGHT
    elif bloco.classificacao == CLASSIF_INVALIDO:
        cell_res.fill = _FILL_RESULT_INVALID
    else:
        cell_res.fill = _FILL_WHITE
        
    if bloco.classificacao == CLASSIF_INVALIDO:
        for r_iter in range(row_ini, r_resultado + 1):
            for c_iter in range(col_ini, c_fim + 1):
                ws.cell(row=r_iter, column=c_iter).fill = _FILL_GRAY


def _escrever_cabecalho(ws, mapa: MapaPlaca, total_cols_excel: int, nome_operador: str) -> int:
    """Escreve cabecalho (nome exame + placa OK) e retorna a linha onde o grid comeca."""
    # Linha 1: nome exame
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols_excel)
    c = ws.cell(row=1, column=1, value=mapa.nome_exame.upper())
    c.font = _font(14, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = _FILL_HEADER_LIGHT
    ws.row_dimensions[1].height = 24

    # Linha 2: placa info
    ws.merge_cells(
        start_row=2, start_column=1, end_row=2, end_column=max(1, total_cols_excel // 2)
    )
    nome_user = nome_operador if nome_operador else "Usuário Desconhecido"
    data_atual = datetime.now().strftime("%d/%m/%Y")
    placa_info = ws.cell(row=2, column=1, value=f"Placa: {mapa.nome_placa}  |  Operador: {nome_user}  |  Data: {data_atual}")
    placa_info.font = _font(10)
    placa_info.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Linha 2 (direita): validacao da placa
    col_validacao_ini = max(1, total_cols_excel // 2) + 1
    ws.merge_cells(
        start_row=2,
        start_column=col_validacao_ini,
        end_row=2,
        end_column=total_cols_excel,
    )
    status_text = "PLACA OK" if mapa.placa_ok else "RECOMECAR PROCESSO"
    c_status = ws.cell(row=2, column=col_validacao_ini, value=status_text)
    c_status.font = _font(11, bold=True, white=mapa.placa_ok)
    c_status.fill = _FILL_BLACK if mapa.placa_ok else _FILL_GRAY
    c_status.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # linha 3 = espaco
    return 4  # grid comeca na linha 4


def _escrever_rodape_controles(
    ws, mapa: MapaPlaca, row_ini: int, total_cols_excel: int
) -> int:
    if not mapa.controles:
        return 57

    cn_ctrl = next((c for c in reversed(mapa.controles) if c.nome == "CN"), None)
    cp_ctrl = next((c for c in reversed(mapa.controles) if c.nome == "CP"), None)

    for linha_alvo, ctrl_obj in [(54, cn_ctrl), (55, cp_ctrl)]:
        if not ctrl_obj:
            continue
            
        try:
            # Remove virgulas se houver
            ct_sanitized = str(ctrl_obj.ct_formatado).replace(",", ".")
            ct_val = round(float(ct_sanitized), 3)
            ct_str = f"{ct_val:.3f}"
        except:
            ct_str = ctrl_obj.ct_formatado or "ND"
            
        texto = f"{ctrl_obj.nome}: CT = {ct_str}"
        
        ws.merge_cells(start_row=linha_alvo, start_column=1, end_row=linha_alvo, end_column=24)
        c_cell = ws.cell(row=linha_alvo, column=1, value=texto)
        c_cell.font = _font(10, bold=True)
        c_cell.alignment = Alignment(horizontal="center", vertical="center")
        c_cell.border = _BLOCK_BORDER

    return 57


def _escrever_assinaturas(ws, row_ini: int, total_cols_excel: int, nome_operador: str) -> None:
    from datetime import datetime
    data_str = datetime.now().strftime("%d/%m/%Y")
    meio = max(1, total_cols_excel // 2)
    
    # Analisado por:
    c1 = ws.cell(row=row_ini, column=1, value=f"Analisado por ({nome_operador}): _________________ ({data_str})")
    c1.font = _font(9)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row_ini, start_column=1, end_row=row_ini, end_column=meio)
    
    # Liberado por:
    c2 = ws.cell(
        row=row_ini, column=meio + 1, value=f"Liberado por: _________________________ ({data_str})"
    )
    c2.font = _font(9)
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(
        start_row=row_ini, start_column=meio + 1, end_row=row_ini, end_column=total_cols_excel
    )


def renderizar_xlsx_grade_fisica(
    df_analise: pd.DataFrame,
    mapa: MapaPlaca,
    caminho_arquivo: str,
    exame: Optional[str] = None,
    pocos_por_amostra: Optional[int] = None,
    nome_operador: str = "Usuário Desconhecido",
) -> str:
    """Gera .xlsx com a mesma grade fisica A-H x 1-12 usada na aba Mapa da Placa."""
    from ui.modules.plate_viewer import COL_LABELS, ROW_LABELS, PlateModel

    wb = Workbook()
    ws = wb.active
    ws.title = "Mapa da Placa"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.35, bottom=0.35)
    ws.print_options = PrintOptions(horizontalCentered=True)

    total_cols_excel = PHYSICAL_TOTAL_COLS
    _escrever_cabecalho(ws, mapa, total_cols_excel, nome_operador)

    ws.cell(row=PHYSICAL_GRID_HEADER_ROW, column=PHYSICAL_GRID_LABEL_COL, value="")
    for col_label in COL_LABELS:
        col_idx = _well_excel_col(col_label)
        cell = ws.cell(row=PHYSICAL_GRID_HEADER_ROW, column=col_idx, value=col_label)
        cell.font = _font(10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = _FILL_HEADER_LIGHT

    for row_label in ROW_LABELS:
        row_idx = _well_excel_row(row_label)
        cell = ws.cell(row=row_idx, column=PHYSICAL_GRID_LABEL_COL, value=row_label)
        cell.font = _font(10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = _FILL_HEADER_LIGHT

    df_model = _normalizar_dataframe_para_plate_model(df_analise)
    plate_model = PlateModel.from_df(
        df_model,
        group_size=pocos_por_amostra,
        exame=exame or mapa.nome_exame,
    )

    blocos_por_codigo = _indexar_blocos_por_codigo(mapa)
    grupos_processados: set[str] = set()

    for row_label in ROW_LABELS:
        for col_label in COL_LABELS:
            well_id = f"{row_label}{int(col_label):02d}"
            well = plate_model.get_well(well_id)

            if well is not None and bool(getattr(well, "is_grouped", False)):
                group_id = str(getattr(well, "group_id", "") or well_id)
                if group_id in grupos_processados:
                    continue
                group_wells = list(getattr(plate_model, "group_dict", {}).get(group_id, [well_id]))
                grupos_processados.add(group_id)
            else:
                group_wells = [well_id]

            valid_wells = []
            for grouped_well in group_wells:
                grouped_well = str(grouped_well or "")
                if len(grouped_well) < 2:
                    continue
                valid_wells.append(grouped_well)
            if not valid_wells:
                valid_wells = [well_id]

            rows = [w[0].upper() for w in valid_wells]
            cols = [str(int(w[1:])) for w in valid_wells]
            row_ini = min(_well_excel_row(r) for r in rows)
            col_ini = min(_well_excel_col(c) for c in cols)
            col_fim = max(_well_excel_col(c) for c in cols)

            representative = well
            if representative is None:
                for grouped_well in valid_wells:
                    representative = plate_model.get_well(grouped_well)
                    if representative is not None:
                        break

            bloco = _bloco_para_well(representative, blocos_por_codigo)
            border = (
                _border_group(int(getattr(representative, "group_size", 1) or 1))
                if representative is not None and bool(getattr(representative, "is_grouped", False))
                else _BLOCK_BORDER
            )
            _escrever_bloco_grade_fisica(ws, row_ini, col_ini, col_fim, bloco, border)

    # Dimensoes estaveis: colunas/linhas de poco com pequenos espacadores em branco.
    ws.column_dimensions[get_column_letter(PHYSICAL_GRID_LABEL_COL)].width = 4
    for col in range(PHYSICAL_GRID_FIRST_COL, total_cols_excel + 1):
        letter = get_column_letter(col)
        if (col - PHYSICAL_GRID_FIRST_COL) % PHYSICAL_WELL_COL_STEP == 0:
            ws.column_dimensions[letter].width = 12.5
        else:
            ws.column_dimensions[letter].width = 2.0

    ws.row_dimensions[PHYSICAL_GRID_HEADER_ROW].height = 18
    last_well_row = _well_excel_row(ROW_LABELS[-1])
    for row in range(PHYSICAL_GRID_FIRST_ROW, last_well_row + PHYSICAL_BLOCK_ROWS):
        if (row - PHYSICAL_GRID_FIRST_ROW) % PHYSICAL_WELL_ROW_STEP == 0:
            ws.row_dimensions[row].height = 16
        elif 1 <= (row - PHYSICAL_GRID_FIRST_ROW) % PHYSICAL_WELL_ROW_STEP <= 3:
            ws.row_dimensions[row].height = 12
        elif (row - PHYSICAL_GRID_FIRST_ROW) % PHYSICAL_WELL_ROW_STEP == 4:
            ws.row_dimensions[row].height = 16
        else:
            ws.row_dimensions[row].height = 6

    rodape_row = last_well_row + PHYSICAL_BLOCK_ROWS + 1
    proxima_linha = _escrever_rodape_controles(ws, mapa, rodape_row, total_cols_excel)
    _escrever_assinaturas(ws, proxima_linha, total_cols_excel, nome_operador)

    diretorio = os.path.dirname(caminho_arquivo)
    if diretorio and not os.path.isdir(diretorio):
        os.makedirs(diretorio, exist_ok=True)

    wb.save(caminho_arquivo)
    return caminho_arquivo


def renderizar_xlsx(mapa: MapaPlaca, caminho_arquivo: str, nome_operador: str = "Usuário Desconhecido") -> str:
    """Gera o arquivo .xlsx em caminho_arquivo. Retorna o caminho final."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapa da Placa"

    # paisagem A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4)
    ws.print_options = PrintOptions(horizontalCentered=True)

    total_cols_excel = mapa.blocos_por_linha * BLOCK_COLS

    # larguras de coluna: blocos pequenos quanto mais blocos por linha
    largura_col = max(7.0, 22.0 / max(1, BLOCK_COLS))  # padrao razoavel
    # ajuste fino conforme densidade: 12 blocos/linha => colunas mais estreitas
    if mapa.blocos_por_linha >= 12:
        largura_col = 6.5
    elif mapa.blocos_por_linha >= 6:
        largura_col = 8.5
    elif mapa.blocos_por_linha >= 4:
        largura_col = 11.0
    else:
        largura_col = 13.5
    for c in range(1, total_cols_excel + 1):
        ws.column_dimensions[get_column_letter(c)].width = largura_col

    # cabecalho
    grid_row_ini = _escrever_cabecalho(ws, mapa, total_cols_excel, nome_operador)

    # grid de blocos
    for i_linha, linha_blocos in enumerate(mapa.blocos):
        r = grid_row_ini + i_linha * BLOCK_ROWS
        for i_col, bloco in enumerate(linha_blocos):
            c_ini = 1 + i_col * BLOCK_COLS
            _escrever_bloco(ws, bloco, r, c_ini)
        # altura uniforme das linhas do bloco
        ws.row_dimensions[r].height = 16  # codigo
        ws.row_dimensions[r + 1].height = 11
        ws.row_dimensions[r + 2].height = 11
        ws.row_dimensions[r + 3].height = 11
        ws.row_dimensions[r + 4].height = 14  # resultado

    grid_row_fim = grid_row_ini + mapa.linhas_blocos * BLOCK_ROWS + 1
    proxima_linha = _escrever_rodape_controles(
        ws, mapa, grid_row_fim, total_cols_excel
    )
    _escrever_assinaturas(ws, proxima_linha, total_cols_excel, nome_operador)

    # garante diretorio
    diretorio = os.path.dirname(caminho_arquivo)
    if diretorio and not os.path.isdir(diretorio):
        os.makedirs(diretorio, exist_ok=True)

    wb.save(caminho_arquivo)
    return caminho_arquivo


# ----------------------------------------------------------------------------
# API publica de alto nivel
# ----------------------------------------------------------------------------


def gerar_mapa_placa_xlsx(
    df_analise: pd.DataFrame,
    nome_exame: str,
    nome_placa: str,
    pocos_por_amostra: int,
    diretorio_saida: str,
    ordem_alvos: Optional[Sequence[str]] = None,
    nome_arquivo: Optional[str] = None,
    nome_operador: str = "Usuário Desconhecido",
) -> str:
    """API de alto nivel: gera o arquivo .xlsx final. Retorna o caminho gerado.

    Args:
        df_analise: DataFrame resultante da analise (com colunas CT_*, Res_*)
        nome_exame: nome amigavel do exame (ex: "ZDC BioManguinhos")
        nome_placa: nome derivado do arquivo de resultados (ex: "placa_680_results_20250423")
        pocos_por_amostra: 1 (VR1e2), 3 (ZDC), etc.
        diretorio_saida: pasta onde salvar (geralmente "relatorios/")
        ordem_alvos: ordem desejada dos alvos no grid; se omitido, infere do DF
        nome_arquivo: nome customizado (sem extensao); se omitido, usa padrao com timestamp

    Returns:
        Caminho absoluto do arquivo gerado.
    """
    mapa = construir_mapa_de_dataframe(
        df_analise=df_analise,
        nome_exame=nome_exame,
        nome_placa=nome_placa,
        pocos_por_amostra=pocos_por_amostra,
        ordem_alvos=ordem_alvos,
    )
    if not nome_arquivo:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        slug_exame = "".join(
            ch if ch.isalnum() else "_" for ch in nome_exame
        ).strip("_")
        slug_placa = "".join(
            ch if ch.isalnum() else "_" for ch in nome_placa
        ).strip("_")
        nome_arquivo = f"mapa_placa_{slug_exame}_{slug_placa}_{ts}"
    caminho = os.path.join(diretorio_saida, f"{nome_arquivo}.xlsx")
    return renderizar_xlsx_grade_fisica(
        df_analise=df_analise,
        mapa=mapa,
        caminho_arquivo=caminho,
        exame=nome_exame,
        pocos_por_amostra=pocos_por_amostra,
        nome_operador=nome_operador,
    )
